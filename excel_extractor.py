#!/usr/bin/env python3
"""
Excel to CSV Extractor with Color-Based Classification
Extracts filtered data from Excel files with color-coded cell processing
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Color
from pathlib import Path
from datetime import datetime
import logging
import sys
from typing import Optional, Dict, List

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('extraction.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class ExcelColorExtractor:
    """Extract and process Excel data with color-based classification"""
    
    # RGB color definitions for blue and red identification
    BLUE_THRESHOLD = {'r': 100, 'g': 100, 'b': 200}
    RED_THRESHOLD = {'r': 200, 'g': 100, 'b': 100}
    
    def __init__(self, source_dir: str, target_sheet: str = "PLANING"):
        """
        Initialize extractor
        
        Args:
            source_dir: Directory containing Excel files
            target_sheet: Name of sheet to process
        """
        self.source_dir = Path(source_dir)
        self.target_sheet = target_sheet
        self.workbook: Optional[openpyxl.Workbook] = None
        self.worksheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        
    def find_target_file(self) -> Optional[Path]:
        """
        Find Excel file containing 'CW' and 'LTP' in filename
        
        Returns:
            Path to target file or None
        """
        logger.info(f"Searching for files in: {self.source_dir}")
        
        if not self.source_dir.exists():
            logger.error(f"Directory does not exist: {self.source_dir}")
            return None
            
        for file_path in self.source_dir.glob("*.xlsx"):
            filename = file_path.name.upper()
            if "CW" in filename and "LTP" in filename:
                logger.info(f"Found target file: {file_path.name}")
                return file_path
                
        logger.warning("No file containing 'CW' and 'LTP' found")
        return None
    
    def load_workbook(self, file_path: Path) -> bool:
        """
        Load Excel workbook and target sheet
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Loading workbook: {file_path}")
            self.workbook = openpyxl.load_workbook(
                file_path,
                data_only=False,
                keep_vba=False
            )
            
            if self.target_sheet not in self.workbook.sheetnames:
                logger.error(f"Sheet '{self.target_sheet}' not found. Available: {self.workbook.sheetnames}")
                return False
                
            self.worksheet = self.workbook[self.target_sheet]
            logger.info(f"Loaded sheet: {self.target_sheet}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return False
    
    def is_blue_cell(self, cell) -> bool:
        """
        Check if cell has blue fill color
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            True if cell is blue
        """
        if not cell.fill or not cell.fill.start_color:
            return False
            
        color = cell.fill.start_color
        
        # Handle indexed colors and RGB
        if hasattr(color, 'rgb') and color.rgb:
            rgb_str = color.rgb
            if len(rgb_str) == 8:  # ARGB format
                rgb_str = rgb_str[2:]
            
            try:
                r = int(rgb_str[0:2], 16)
                g = int(rgb_str[2:4], 16)
                b = int(rgb_str[4:6], 16)
                
                # Blue detection: high blue value, lower red/green
                return (b > self.BLUE_THRESHOLD['b'] and 
                        r < self.BLUE_THRESHOLD['r'] and 
                        g < self.BLUE_THRESHOLD['g'])
            except:
                return False
                
        return False
    
    def is_red_cell(self, cell) -> bool:
        """
        Check if cell has red fill color
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            True if cell is red
        """
        if not cell.fill or not cell.fill.start_color:
            return False
            
        color = cell.fill.start_color
        
        if hasattr(color, 'rgb') and color.rgb:
            rgb_str = color.rgb
            if len(rgb_str) == 8:  # ARGB format
                rgb_str = rgb_str[2:]
            
            try:
                r = int(rgb_str[0:2], 16)
                g = int(rgb_str[2:4], 16)
                b = int(rgb_str[4:6], 16)
                
                # Red detection: high red value, lower green/blue
                return (r > self.RED_THRESHOLD['r'] and 
                        g < self.RED_THRESHOLD['g'] and 
                        b < self.RED_THRESHOLD['b'])
            except:
                return False
                
        return False
    
    def get_cell_type(self, cell) -> str:
        """
        Determine cell type based on color
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            'ASSY' for blue, 'SEW' for red, '' for others
        """
        if self.is_blue_cell(cell):
            return 'ASSY'
        elif self.is_red_cell(cell):
            return 'SEW'
        return ''
    
    def extract_data(self) -> pd.DataFrame:
        """
        Extract filtered data from worksheet
        
        Returns:
            DataFrame with extracted and processed data
        """
        logger.info("Starting data extraction")
        
        data_rows = []
        header_row = None
        date_row = 61  # Row with dates (DD/MM format)
        
        # Find header row and extract column positions
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, max_row=100), start=1):
            cell_d = row[3] if len(row) > 3 else None  # Column D
            if cell_d and cell_d.value == "Week":
                header_row = row_idx
                logger.info(f"Found header at row {header_row}")
                break
        
        if not header_row:
            logger.error("Could not find header row with 'Week' in column D")
            return pd.DataFrame()
        
        # Extract headers
        headers = []
        for cell in self.worksheet[header_row]:
            headers.append(cell.value if cell.value else f"Col_{cell.column}")
        headers.append("Type")  # Additional column for ASSY/SEW
        
        # Extract dates from row 61
        date_mapping = {}
        for cell in self.worksheet[date_row]:
            if cell.value and cell.column > 6:  # Start from column G (7)
                date_mapping[cell.column] = cell.value
        
        logger.info(f"Found {len(date_mapping)} date columns starting from row {date_row}")
        
        # Process data rows
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=header_row+1), start=header_row+1):
            if row_idx == date_row:  # Skip date row
                continue
                
            cell_d = row[3] if len(row) > 3 else None  # Column D (index 3)
            
            if cell_d and cell_d.value == "Weekly Output PLAN":
                row_data = []
                row_type = ""
                
                # Extract all cell values
                for cell in row:
                    row_data.append(cell.value)
                    
                    # Check color for type determination (columns after F)
                    if cell.column > 6 and not row_type:
                        cell_type = self.get_cell_type(cell)
                        if cell_type:
                            row_type = cell_type
                
                row_data.append(row_type)  # Add type column
                data_rows.append(row_data)
                logger.debug(f"Extracted row {row_idx}: Type={row_type}")
        
        logger.info(f"Extracted {len(data_rows)} data rows")
        
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        return df
    
    def save_to_csv(self, df: pd.DataFrame, output_path: str = None) -> bool:
        """
        Save DataFrame to CSV file
        
        Args:
            df: DataFrame to save
            output_path: Output file path (optional)
            
        Returns:
            True if successful
        """
        if df.empty:
            logger.warning("No data to save")
            return False
        
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"extracted_data_{timestamp}.csv"
        
        try:
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            logger.info(f"Data saved to: {output_path}")
            logger.info(f"Total rows: {len(df)}, Total columns: {len(df.columns)}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to save CSV: {e}")
            return False
    
    def process(self, output_path: str = None) -> bool:
        """
        Execute complete extraction process
        
        Args:
            output_path: Optional output file path
            
        Returns:
            True if successful
        """
        try:
            # Find target file
            file_path = self.find_target_file()
            if not file_path:
                return False
            
            # Load workbook
            if not self.load_workbook(file_path):
                return False
            
            # Extract data
            df = self.extract_data()
            
            # Save to CSV
            return self.save_to_csv(df, output_path)
            
        finally:
            if self.workbook:
                self.workbook.close()


def main():
    """Main execution function"""
    
    # Configuration
    SOURCE_DIR = r"G:\All\Long-term planning"
    TARGET_SHEET = "PLANING"
    OUTPUT_FILE = "filtered_output.csv"
    
    # Initialize and process
    extractor = ExcelColorExtractor(SOURCE_DIR, TARGET_SHEET)
    
    logger.info("="*60)
    logger.info("Excel Color-Based Extraction Tool")
    logger.info("="*60)
    
    success = extractor.process(OUTPUT_FILE)
    
    if success:
        logger.info("✓ Extraction completed successfully")
        return 0
    else:
        logger.error("✗ Extraction failed")
        return 1


if __name__ == "__main__":
    sys.exit(main())
