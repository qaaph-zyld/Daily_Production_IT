import os
import csv
import json
import re
import pyodbc
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.utils import column_index_from_string
try:
    from openpyxl.styles.colors import COLOR_INDEX
except Exception:
    COLOR_INDEX = None
import calendar
from flask import Flask, render_template, jsonify
from flask_cors import CORS
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
from decimal import Decimal
try:
    import win32com.client as win32
    import pythoncom
except Exception:
    win32 = None
    pythoncom = None

# Load .env
load_dotenv()

# Load settings.json if available (optional config override)
SETTINGS = {}
SETTINGS_PATH = os.path.join(os.path.dirname(__file__), 'config', 'settings.json')
if os.path.exists(SETTINGS_PATH):
    try:
        with open(SETTINGS_PATH, 'r', encoding='utf-8') as f:
            SETTINGS = json.load(f)
        print(f"[CONFIG] Loaded settings from {SETTINGS_PATH}")
    except Exception as e:
        print(f"[CONFIG] Warning: Could not load settings.json: {e}")

app = Flask(__name__)
CORS(app)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.jinja_env.cache = {}

# DB config
DB_CONFIG = {
    'server': os.getenv('DB_SERVER', 'a265m001'),
    'database': os.getenv('DB_DATABASE', 'QADEE2798'),
    'username': os.getenv('DB_USERNAME', 'PowerBI'),
    'password': os.getenv('DB_PASSWORD', 'P0werB1'),
}

# Server config
FLASK_HOST = os.getenv('FLASK_HOST', '0.0.0.0')
PVS_PORT = int(os.getenv('PVS_PORT', '5051'))

# PVS config
_BASE_DIR = os.path.dirname(__file__)
_planned_default = os.path.join('PVS', 'Planned_qtys.xlsx')
_map_default = os.path.join('PVS', 'ProdLine_Project_Map.csv')

PVS_PLANNED_XLSX = os.getenv('PVS_PLANNED_XLSX', _planned_default)
if not os.path.isabs(PVS_PLANNED_XLSX):
    PVS_PLANNED_XLSX = os.path.join(_BASE_DIR, PVS_PLANNED_XLSX)

PVS_RECALC_XLSX = os.getenv('PVS_RECALC_XLSX', 'false').strip().lower() in ('1', 'true', 'yes')
PVS_ADHERENCE_CLAMP = float(os.getenv('PVS_ADHERENCE_CLAMP', '300'))  # percent cap; values beyond are treated as 0%

PVS_MAP_CSV = os.getenv('PVS_MAP_CSV', _map_default)
if not os.path.isabs(PVS_MAP_CSV):
    PVS_MAP_CSV = os.path.join(_BASE_DIR, PVS_MAP_CSV)
if not os.path.exists(PVS_MAP_CSV):
    _alt_map = os.path.join(_BASE_DIR, 'dist', 'PVS', 'ProdLine_Project_Map.csv')
    if os.path.exists(_alt_map):
        PVS_MAP_CSV = _alt_map

# External WH Receipt workbook (used to avoid Excel on VM)
_DATA_SOURCES = SETTINGS.get('dataSources', {}) if isinstance(SETTINGS, dict) else {}
PVS_EXTERNAL_XLSX = _DATA_SOURCES.get(
    'externalSourceExcel',
    r"G:\Logistics\6_Reporting\1_PVS\WH Receipt FY25.xlsx",
)
PVS_EXTERNAL_SHEET = _DATA_SOURCES.get('externalSheetName', 'Daily PVS')
PVS_EXTERNAL_TARGET_LABEL = _DATA_SOURCES.get('externalTargetLabel', 'Target (LTP input)')

PVS_LTP_DIR = _DATA_SOURCES.get('ltpDirectory', r"G:\All\Long-term planning")
PVS_LTP_SHEET = _DATA_SOURCES.get('ltpSheetName', 'Planning')
PVS_LTP_REF_CSV = _DATA_SOURCES.get('ltpRefCsv', os.path.join(_BASE_DIR, 'ref.csv'))
if not os.path.isabs(PVS_LTP_REF_CSV):
    PVS_LTP_REF_CSV = os.path.join(_BASE_DIR, PVS_LTP_REF_CSV)
_alt_ref = os.path.join(_BASE_DIR, 'PVS', 'ref.csv')
if os.path.exists(_alt_ref):
    PVS_LTP_REF_CSV = _alt_ref
elif not os.path.exists(PVS_LTP_REF_CSV):
    if os.path.exists(_alt_ref):
        PVS_LTP_REF_CSV = _alt_ref
PVS_LTP_KEYWORDS = _DATA_SOURCES.get('ltpFilenameKeywords', ['FY', 'CW', 'LTP'])
PVS_LTP_DATE_ROW = int(_DATA_SOURCES.get('ltpDateRow', 63))
PVS_LTP_DATE_START_COL = _DATA_SOURCES.get('ltpDateStartCol', 'X')
PVS_LTP_DATE_END_COL = _DATA_SOURCES.get('ltpDateEndCol', 'BW')
PVS_LTP_LABEL = _DATA_SOURCES.get('ltpLabel', 'Weekly Output Plan')
PVS_LTP_FALLBACK_CSV = _DATA_SOURCES.get('ltpFallbackCsv', os.path.join(_BASE_DIR, 'LTP_extracted.csv'))
if not os.path.isabs(PVS_LTP_FALLBACK_CSV):
    PVS_LTP_FALLBACK_CSV = os.path.join(_BASE_DIR, PVS_LTP_FALLBACK_CSV)
PVS_LTP_FORMULAS_XLSX = _DATA_SOURCES.get('ltpFormulasWorkbook', 'LTP_formulas.xlsx')
if PVS_LTP_FORMULAS_XLSX and not os.path.isabs(PVS_LTP_FORMULAS_XLSX):
    PVS_LTP_FORMULAS_XLSX = os.path.join(_BASE_DIR, PVS_LTP_FORMULAS_XLSX)
PVS_LTP_FORMULAS_SHEET = _DATA_SOURCES.get('ltpFormulasSheetName', 'LTP_Formulas')

# Monthly OLK workbook (for monthly OLK targets by prod line)
PVS_OLK_XLSX = _DATA_SOURCES.get(
    'monthlyOlkExcel',
    r"G:\Logistics\6_Reporting\1_PVS\Monthly_OLK.xlsx",
)

# OLK CSV (preferred source — PVS/OLK.csv)
_olk_csv_default = os.path.join('PVS', 'OLK.csv')
PVS_OLK_CSV = _DATA_SOURCES.get('olkCsv', _olk_csv_default)
if PVS_OLK_CSV and not os.path.isabs(PVS_OLK_CSV):
    PVS_OLK_CSV = os.path.join(_BASE_DIR, PVS_OLK_CSV)

_BEHAVIOR = SETTINGS.get('behavior', {}) if isinstance(SETTINGS, dict) else {}
_PLAN_SOURCE = str(_BEHAVIOR.get('planSource', '') or '').strip().lower()
PVS_PLAN_SOURCE = _PLAN_SOURCE or ('wh_receipt' if _BEHAVIOR.get('useWhReceiptForPlan', True) else 'planned_xlsx')
PVS_USE_WH_RECEIPT = PVS_PLAN_SOURCE == 'wh_receipt'
PVS_LTP_WORKDAYS_PER_WEEK = int(_BEHAVIOR.get('ltpWorkdaysPerWeek', 5) or 5)
PVS_SHOW_WEEKEND_ON_MONDAY = bool(_BEHAVIOR.get('showWeekendDataOnMonday', True))
PVS_REGENERATE_INPUTS = bool(_BEHAVIOR.get('regenerateInputsOnCompute', True))
PVS_EXPORT_LTP_REF_EXTRACT = bool(_BEHAVIOR.get('exportLtpRefExtractOnCompute', True))
PVS_LTP_REF_EXTRACT_CSV = _DATA_SOURCES.get('ltpRefExtractCsv', os.path.join('PVS', 'Debug', 'LTP_ref_extract.csv'))
if PVS_LTP_REF_EXTRACT_CSV and not os.path.isabs(PVS_LTP_REF_EXTRACT_CSV):
    PVS_LTP_REF_EXTRACT_CSV = os.path.join(_BASE_DIR, PVS_LTP_REF_EXTRACT_CSV)
PVS_PROD_SQL_PATH = _DATA_SOURCES.get('productionSql', os.path.join('PVS', 'Production', 'PVS-Production.sql'))
if PVS_PROD_SQL_PATH and not os.path.isabs(PVS_PROD_SQL_PATH):
    PVS_PROD_SQL_PATH = os.path.join(_BASE_DIR, PVS_PROD_SQL_PATH)
PVS_PROD_TR_TYPES = [
    str(t).strip().upper()
    for t in (_BEHAVIOR.get('productionTrTypes') or ['RCT-WO'])
    if str(t).strip()
]

SEW_NAME_OVERRIDES = {
    'BJA',
    'MAN',
    'PIP',
    'PZ1D',
    'PRE PRODUCTION',
    'RENAULT',
    'SCANIA',
}

GROUP_OVERRIDES = {
    # Commercial Vehicles (CV)
    'RENAULT': 'CV',
    'MAN': 'CV',
    'SCANIA': 'CV',
    # PIP project (P13A SEW line maps to PIP project)
    'P13A': 'PIP',
    # Volvo (combine SEW and ASSY into one group)
    'VOLVO- SEW': 'VOLVO',
    'VOLVO- ASSY': 'VOLVO',
    'VOLVO SEW': 'VOLVO',
    'VOLVO ASSY': 'VOLVO',
}


def norm_code(code: str) -> str:
    s = (code or '')
    s = s.replace('\xa0', ' ')
    s = s.strip().upper()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('-', '_').replace(' ', '_')
    s = re.sub(r'_+', '_', s)
    return s


def load_map_csv(path: str) -> dict[str, str]:
    """Return dict: {prod_line_code -> project_name} with uppercase codes."""
    mapping: dict[str, str] = {}
    if not os.path.exists(path):
        return mapping
    with open(path, 'r', newline='', encoding='utf-8') as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            code = norm_code(row.get('prod_line', ''))
            proj = (row.get('project') or '').strip()
            if code and proj:
                mapping[code] = proj
    return mapping


def load_olk_csv(path: str) -> dict[str, float]:
    """Return dict: {label -> monthly_olk_qty} from PVS/OLK.csv.

    CSV format:  Dashboard page,<MonthName>
                 <page_num>,<label>,<qty>
    Labels are dashboard display names (e.g. BJA, CDPO - SEW, Volvo - ASSY).
    """
    data: dict[str, float] = {}
    if not path or not os.path.exists(path):
        print(f"[OLK] OLK CSV not found: {path}")
        return data

    try:
        with open(path, 'r', newline='', encoding='utf-8-sig') as f:
            first_line = f.readline()
            for raw in f:
                line = (raw or '').strip()
                if not line:
                    continue
                parts = [p.strip() for p in line.split(',')]
                if len(parts) < 3:
                    continue
                label = parts[1].strip()
                if not label:
                    continue
                try:
                    val = float(parts[2] or 0)
                except (ValueError, TypeError):
                    val = 0.0
                if val:
                    data[label] = max(data.get(label, 0.0), val)
    except Exception as e:
        print(f"[OLK] ERROR reading OLK CSV: {e}")
        return data

    print(f"[OLK] Loaded {len(data)} OLK rows from CSV: {path}")
    return data


def load_monthly_olk(path: str) -> dict[str, float]:
    """Return dict: {prod_line_code -> monthly_olk_qty} from Monthly_OLK.xlsx.

    Reads the "Monthly OLK" sheet and uses columns "Row Labels" (prod line
    codes such as B_FG, Q_FG, etc.) and "Sum of Monthly OLK" for the qty.
    """
    data: dict[str, float] = {}
    if not path or not os.path.exists(path):
        print(f"[OLK] Monthly OLK workbook not found: {path}")
        return data

    try:
        df = pd.read_excel(path, sheet_name='Monthly OLK')
    except Exception as e:
        print(f"[OLK] ERROR opening Monthly OLK workbook: {e}")
        return data

    if df.empty:
        print("[OLK] Monthly OLK sheet is empty")
        return data

    col_code = None
    col_val = None
    for col in df.columns:
        name = str(col).strip().lower()
        if name == 'row labels':
            col_code = col
        elif name == 'sum of monthly olk':
            col_val = col

    if col_code is None or col_val is None:
        print("[OLK] Required columns 'Row Labels' / 'Sum of Monthly OLK' not found")
        return data

    for _, row in df.iterrows():
        code = norm_code(str(row.get(col_code, '')))
        if not code:
            continue
        try:
            val = float(row.get(col_val, 0) or 0.0)
        except Exception:
            val = 0.0
        if not val:
            continue
        # If pivot repeats codes, accumulate
        data[code] = data.get(code, 0.0) + val

    print(f"[OLK] Loaded {len(data)} OLK rows from Monthly OLK workbook")
    return data


def _choose_sql_driver():
    try:
        drivers = list(pyodbc.drivers())
    except Exception:
        drivers = []
    for name in (
        'ODBC Driver 18 for SQL Server',
        'ODBC Driver 17 for SQL Server',
        'SQL Server',
    ):
        if name in drivers:
            return name
    return 'SQL Server'


def get_db_connection():
    driver = _choose_sql_driver()
    extra = 'TrustServerCertificate=yes;'
    if 'ODBC Driver 18' in driver:
        extra = 'Encrypt=no;TrustServerCertificate=yes;'
    conn_str = (
        f"DRIVER={{{driver}}};SERVER={DB_CONFIG['server']};DATABASE={DB_CONFIG['database']};"
        f"UID={DB_CONFIG['username']};PWD={DB_CONFIG['password']};{extra}"
    )
    return pyodbc.connect(conn_str)


_TR_HIST_COLS: set[str] | None = None


def _get_tr_hist_columns(conn) -> set[str]:
    global _TR_HIST_COLS
    if _TR_HIST_COLS is not None:
        return _TR_HIST_COLS
    cols: set[str] = set()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME='tr_hist'"
        )
        for (name,) in cur.fetchall():
            if name:
                cols.add(str(name).strip().lower())
        cur.close()
    except Exception:
        cols = set()
    _TR_HIST_COLS = cols
    return cols


def _parse_date_ddmmyyyy(s: str) -> date | None:
    s = s.strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, '%d/%m/%Y').date()
    except Exception:
        return None


def _coerce_header_to_date(col) -> date | None:
    try:
        # Skip NaN/None early
        if col is None or (isinstance(col, float) and pd.isna(col)):
            return None
        # Direct types
        if isinstance(col, date) and not isinstance(col, datetime):
            return col
        if isinstance(col, datetime):
            return col.date()
        # Pandas Timestamp
        if 'Timestamp' in type(col).__name__:
            try:
                if pd.isna(col):
                    return None
                return col.date()
            except Exception:
                pass
        # Excel serial number (float/int) - must be a reasonable date range
        if isinstance(col, (int, float)):
            # Excel dates for 2020-2030 are roughly 43831 to 47848
            if 40000 < col < 50000:
                try:
                    d = pd.to_datetime(col, unit='D', origin='1899-12-30', errors='raise')
                    return d.date()
                except Exception:
                    pass
            return None
        # String formats: dd/mm/yyyy, ISO, or generic
        s = str(col).strip()
        d = _parse_date_ddmmyyyy(s)
        if d:
            return d
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            pass
        try:
            d2 = pd.to_datetime(s, dayfirst=True, errors='coerce')
            if pd.notna(d2):
                return d2.date()
        except Exception:
            pass
    except Exception:
        return None
    return None


def _split_sql_batches(sql_text: str) -> list[str]:
    if not sql_text:
        return []
    lines = sql_text.splitlines()
    batches: list[str] = []
    cur: list[str] = []
    for raw in lines:
        line = raw.rstrip('\n')
        if re.match(r'^\s*GO\s*$', line, flags=re.IGNORECASE):
            batch = '\n'.join(cur).strip()
            if batch:
                batches.append(batch)
            cur = []
            continue
        cur.append(line)
    last = '\n'.join(cur).strip()
    if last:
        batches.append(last)
    return batches


def _run_production_sql_and_overwrite_csvs(sql_path: str) -> bool:
    if not sql_path or not os.path.exists(sql_path):
        print(f"[SQL] Production SQL not found: {sql_path}")
        return False

    try:
        with open(sql_path, 'r', encoding='utf-8-sig') as f:
            sql_text = f.read()
    except Exception as e:
        print(f"[SQL] ERROR reading production SQL: {e}")
        return False

    batches = _split_sql_batches(sql_text)
    if not batches:
        print("[SQL] No executable SQL batches found")
        return False

    prod_dir = os.path.join(_BASE_DIR, 'PVS', 'Production')
    os.makedirs(prod_dir, exist_ok=True)
    out_paths = [
        os.path.join(prod_dir, '1_PVS_per_Project.csv'),
        os.path.join(prod_dir, '2_PVS_per_SEW.csv'),
        os.path.join(prod_dir, '3_PVS_per_ASSY.csv'),
    ]
    master = _load_master_list(os.path.join(_BASE_DIR, 'PVS', 'master_list.csv'))
    master_by_rs = [master.get('PROJECT') or [], master.get('SEW') or [], master.get('ASSY') or []]

    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        for b in batches[:-1]:
            try:
                cur.execute(b)
                while True:
                    try:
                        more = cur.nextset()
                    except Exception:
                        more = False
                    if not more:
                        break
            except Exception as e:
                print(f"[SQL] WARNING executing batch: {e}")

        cur.execute(batches[-1])
        rs_idx = 0
        while True:
            cols = [c[0] for c in (cur.description or [])]
            rows = cur.fetchall() if cols else []
            if cols and rs_idx < len(out_paths):
                out_path = out_paths[rs_idx]
                day_count = max(len(cols) - 1, 0)
                by_norm: dict[str, tuple[str, list[str]]] = {}
                for r in rows:
                    label = str(r[0] or '').strip()
                    if not label:
                        continue
                    vals: list[str] = []
                    for v in r[1:]:
                        try:
                            vals.append(f"{float(v or 0):.2f}")
                        except Exception:
                            vals.append("0.00")
                    if len(vals) < day_count:
                        vals.extend(["0.00"] * (day_count - len(vals)))
                    by_norm[_norm_key(label)] = (label, vals)

                padded: list[tuple[str, list[str]]] = []
                canon = master_by_rs[rs_idx] if rs_idx < len(master_by_rs) else []
                if canon:
                    for c in canon:
                        k = _norm_key(c)
                        if k in by_norm:
                            padded.append(by_norm[k])
                        else:
                            padded.append((c, ["0.00"] * day_count))
                    # append extras at end
                    for k, (lbl, vals) in by_norm.items():
                        if k not in {_norm_key(c) for c in canon}:
                            padded.append((lbl, vals))
                else:
                    padded = [v for v in by_norm.values()]

                with open(out_path, 'w', newline='', encoding='utf-8') as f:
                    for lbl, vals in padded:
                        if not lbl:
                            continue
                        f.write(','.join([lbl] + vals) + "\n")
                    f.write("\n")
                print(f"[SQL] Wrote {out_path} ({len(padded)} rows)")
                rs_idx += 1

            try:
                has_next = cur.nextset()
            except Exception:
                has_next = False
            if not has_next:
                break

        cur.close()
        try:
            conn.commit()
        except Exception:
            pass
        if rs_idx < 3:
            print(f"[SQL] WARNING: expected 3 result sets, got {rs_idx}")
        return rs_idx >= 1
    except Exception as e:
        print(f"[SQL] ERROR executing production SQL: {e}")
        return False
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass


def _export_ltp_ref_extract_csv(
    out_path: str,
    directory: str,
    sheet_name: str,
    label: str,
    ref_csv: str,
    keywords,
    date_row: int,
    date_start_col: str,
    date_end_col: str,
) -> bool:
    if not out_path:
        return False
    if not ref_csv or not os.path.exists(ref_csv):
        print(f"[LTP-EXTRACT] Reference CSV not found: {ref_csv}")
        return False

    workbook_path = _find_ltp_workbook(directory, keywords)
    if not workbook_path:
        print("[LTP-EXTRACT] No workbook found; skipping")
        return False

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)
    except Exception as e:
        print(f"[LTP-EXTRACT] ERROR opening workbook {workbook_path}: {e}")
        return False

    try:
        desired = (sheet_name or '').strip()
        candidates: list[str] = []
        if desired:
            candidates.append(desired)
        candidates.extend(['Planned', 'Planning', 'PLANING'])
        resolved_sheet = ''
        for cand in candidates:
            if cand in wb.sheetnames:
                resolved_sheet = cand
                break
            lower_match = next((s for s in wb.sheetnames if s.lower() == cand.lower()), '')
            if lower_match:
                resolved_sheet = lower_match
                break
        if not resolved_sheet:
            print(f"[LTP-EXTRACT] Sheet '{sheet_name}' not found in {workbook_path.name}")
            return False

        ws = wb[resolved_sheet]
        start_col_idx = column_index_from_string(str(date_start_col).strip() or 'X')
        end_col_idx = column_index_from_string(str(date_end_col).strip() or 'BW')
        if end_col_idx < start_col_idx:
            start_col_idx, end_col_idx = end_col_idx, start_col_idx

        date_headers: list[date | None] = []
        for col in range(start_col_idx, end_col_idx + 1):
            date_headers.append(_coerce_header_to_date(ws.cell(row=date_row, column=col).value))

        if sum(1 for d in date_headers if d) < 3:
            scan_max = min(ws.max_row, 200)
            best_row = None
            best_count = 0
            for r in range(1, scan_max + 1):
                cnt = 0
                for col in range(start_col_idx, end_col_idx + 1):
                    if _coerce_header_to_date(ws.cell(row=r, column=col).value):
                        cnt += 1
                if cnt > best_count:
                    best_count = cnt
                    best_row = r
            if best_row is not None and best_count >= 3:
                print(f"[LTP-EXTRACT] Date header row auto-detected: {best_row} ({best_count} date-like columns)")
                date_row = best_row
                date_headers = []
                for col in range(start_col_idx, end_col_idx + 1):
                    date_headers.append(_coerce_header_to_date(ws.cell(row=date_row, column=col).value))

        date_cols: list[tuple[int, date]] = []
        for offset, col in enumerate(range(start_col_idx, end_col_idx + 1)):
            d = date_headers[offset] if offset < len(date_headers) else None
            if d:
                date_cols.append((col, d))
        if not date_cols:
            print("[LTP-EXTRACT] No date columns detected; skipping")
            return False

        target_label = (label or '').strip().lower()
        label_col_idx: int | None = 4 if target_label else None
        if target_label:
            scan_rows = min(ws.max_row, 500)
            direct_hits = 0
            for r in range(1, scan_rows + 1):
                v = ws.cell(row=r, column=4).value
                if str(v or '').strip().lower() == target_label:
                    direct_hits += 1
            if direct_hits == 0:
                best_col = None
                best_hits = 0
                scan_cols = min(ws.max_column, 15)
                for c in range(1, scan_cols + 1):
                    hits = 0
                    for r in range(1, scan_rows + 1):
                        v = ws.cell(row=r, column=c).value
                        if str(v or '').strip().lower() == target_label:
                            hits += 1
                    if hits > best_hits:
                        best_hits = hits
                        best_col = c
                if best_col is not None and best_hits > 0:
                    label_col_idx = best_col
                    print(f"[LTP-EXTRACT] Label column auto-detected: {label_col_idx} ({best_hits} matches)")
                else:
                    label_col_idx = None
                    print(f"[LTP-EXTRACT] ERROR: Label '{label}' not found; export skipped")
                    return False

        row_lookup: dict[tuple[str, str, str], int] = {}
        for row_idx in range(1, ws.max_row + 1):
            if label_col_idx is not None:
                label_raw = ws.cell(row=row_idx, column=label_col_idx).value
                if (str(label_raw or '').strip().lower()) != target_label:
                    continue
            project_key = _normalize_ltp_key(ws.cell(row=row_idx, column=1).value)
            model_key = _normalize_ltp_key(ws.cell(row=row_idx, column=2).value)
            row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=3).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=5).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=6).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=4).value)
            if not row_type:
                row_type = _normalize_ltp_type(project_key) or _normalize_ltp_type(model_key)
            if not row_type:
                row_type = _infer_type_from_row(ws, row_idx, start_col_idx, end_col_idx)
            if not row_type:
                row_type = _infer_type_from_row(ws, row_idx, 1, max(end_col_idx, 10))
            if not row_type:
                row_type = _infer_type_from_context(ws, row_idx)
            if not project_key or not model_key or not row_type:
                continue
            k = (project_key, model_key, row_type)
            if k not in row_lookup:
                row_lookup[k] = row_idx

        meta_rows: list[dict[str, str]] = []
        with open(ref_csv, 'r', newline='', encoding='utf-8-sig') as f:
            rdr = csv.DictReader(f)
            for r in rdr:
                meta_rows.append({k: (str(v or '').strip()) for k, v in (r or {}).items()})

        out_dir = os.path.dirname(out_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        base_headers = ['PROJECT', 'SEW', 'ASSY', 'Production Line', 'Model', 'SEW/ASSY', 'Project_Key']
        date_headers_out = [d.isoformat() for _, d in date_cols]
        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            f.write(','.join(base_headers + date_headers_out) + "\n")
            written = 0
            missing = 0
            missing_keys: list[str] = []
            for mr in meta_rows:
                ltp_project = _normalize_ltp_key(
                    mr.get('Production Line')
                    or mr.get('production line')
                    or mr.get('Project')
                    or mr.get('project')
                    or ''
                )
                ltp_model = _normalize_ltp_key(mr.get('Model') or mr.get('model') or '')
                ltp_type = _normalize_ltp_type(mr.get('SEW/ASSY') or mr.get('sew/assy') or mr.get('type') or '')
                excel_row = row_lookup.get((ltp_project, ltp_model, ltp_type))
                if excel_row is None:
                    missing += 1
                    missing_keys.append(f"{ltp_project}/{ltp_model}/{ltp_type}")
                out_row: list[str] = []
                for h in base_headers:
                    out_row.append(str(mr.get(h) or ''))
                if excel_row is None:
                    out_row.extend(['0' for _ in date_cols])
                else:
                    for col, _d in date_cols:
                        val = ws.cell(row=excel_row, column=col).value
                        try:
                            if val is None or (isinstance(val, float) and pd.isna(val)):
                                out_row.append('0')
                            else:
                                out_row.append(str(int(round(float(val)))))
                        except Exception:
                            out_row.append('0')
                f.write(','.join(out_row) + "\n")
                written += 1

        if missing:
            print(f"[LTP-EXTRACT] Missing mappings ({missing}):")
            for k in missing_keys:
                print(f"  - {k}")
        else:
            print(f"[LTP-EXTRACT] Wrote {out_path} ({written} ref rows, 0 missing mappings)")
        return True
    finally:
        try:
            wb.close()
        except Exception:
            pass


def recalc_excel_workbook(path: str) -> bool:
    if not path or not os.path.exists(path):
        return False
    try:
        if win32 is None or pythoncom is None:
            return False
        pythoncom.CoInitialize()
        xl = win32.DispatchEx('Excel.Application')
        try:
            xl.Visible = False
            xl.DisplayAlerts = False
            try:
                xl.AskToUpdateLinks = False
            except Exception:
                pass
            wb = xl.Workbooks.Open(os.path.abspath(path), UpdateLinks=3)
            try:
                try:
                    links = wb.LinkSources(1)  # 1 = xlLinkTypeExcelLinks
                    if links:
                        for ln in links:
                            try:
                                wb.UpdateLink(ln, 1)
                            except Exception:
                                pass
                except Exception:
                    pass
                try:
                    wb.RefreshAll()
                except Exception:
                    pass
                try:
                    xl.CalculateUntilAsyncQueriesDone()
                except Exception:
                    pass
                try:
                    xl.CalculateFullRebuild()
                except Exception:
                    try:
                        xl.CalculateFull()
                    except Exception:
                        pass
                try:
                    wb.Save()
                except Exception:
                    pass
            finally:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    pass
        finally:
            try:
                xl.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        return True
    except Exception:
        return False


def load_planned_from_wh_receipt(path: str, sheet: str, target_label: str):
    """Build planned schedule directly from WH Receipt workbook.

    Reads the 'Daily PVS' (or configured) sheet and uses rows where the
    classification column equals target_label (e.g. 'Target (LTP input)').

    Returns dict: {prod_line_code: {date: qty_int}} where prod_line_code is the
    production line code (e.g. 'B_FG', 'H_FG', ...).
    """
    planned: dict[str, dict[date, int]] = {}
    if not path or not os.path.exists(path):
        print(f"[LOAD-WH] ERROR: External XLSX not found: {path}")
        return planned

    print(f"[LOAD-WH] Loading WH Receipt data from: {path} (sheet={sheet})")

    try:
        df = pd.read_excel(path, sheet_name=sheet, header=None)
    except Exception as e:
        print(f"[LOAD-WH] ERROR opening workbook: {e}")
        return planned

    if df.empty:
        print("[LOAD-WH] Sheet is empty")
        return planned

    # 1) Detect header row that contains dates across many columns
    best_row = None
    best_count = 0
    max_check_rows = min(15, df.shape[0])
    for r in range(max_check_rows):
        row = df.iloc[r, :]
        cnt = 0
        for v in row:
            if _coerce_header_to_date(v) is not None:
                cnt += 1
        if cnt > best_count:
            best_count = cnt
            best_row = r

    if best_row is None or best_count < 3:
        print(f"[LOAD-WH] Could not detect date header row (best_row={best_row}, count={best_count})")
        return planned

    print(f"[LOAD-WH] Using row {best_row} as date header with {best_count} date-like columns")
    date_headers = df.iloc[best_row, :]
    date_cols: list[date | None] = []
    for v in date_headers:
        date_cols.append(_coerce_header_to_date(v))

    # 2) For each row where classification column == target_label, collect per-day values
    # We observed layout: col0=Project, col1=Prod line, col2=label (e.g. 'Target (LTP input)')
    label_col_idx = 2
    prod_line_col_idx = 1

    for r in range(df.shape[0]):
        try:
            label_raw = df.iat[r, label_col_idx]
        except Exception:
            continue
        if str(label_raw).strip() != target_label:
            continue

        prod_line_raw = df.iat[r, prod_line_col_idx] if prod_line_col_idx < df.shape[1] else None
        code = norm_code(str(prod_line_raw))
        if not code:
            continue

        per_day: dict[date, int] = {}
        row = df.iloc[r, :]
        for j, d in enumerate(date_cols):
            if d is None:
                continue
            if j >= len(row):
                continue
            val = row.iat[j]
            try:
                if pd.isna(val):
                    q = 0
                else:
                    q = int(round(float(val)))
            except Exception:
                q = 0
            if q:
                per_day[d] = per_day.get(d, 0) + q

        planned[code] = per_day

    print(f"[LOAD-WH] Loaded {len(planned)} production lines from WH Receipt workbook")
    return planned


def load_planned_xlsx(path: str):
    """Return dict: {line_code: {date: qty_int}}. Reads Excel file with formulas."""
    planned: dict[str, dict[date, int]] = {}
    if not os.path.exists(path):
        print(f"[LOAD] ERROR: Planned XLSX not found: {path}")
        return planned
    
    print(f"[LOAD] Loading planned data from: {path}")
    
    try:
        # Read Excel file - pandas will evaluate formulas and return calculated values
        df = pd.read_excel(path, engine='openpyxl')
        
        # First column is 'Project' (line codes), rest are dates
        if df.empty or len(df.columns) < 2:
            return planned
        
        # Parse date columns (skip first column which is 'Project')
        date_cols = []
        for col in df.columns[1:]:
            d = _coerce_header_to_date(col)
            date_cols.append(d)
        
        # Process each row
        for idx, row in df.iterrows():
            code = norm_code(str(row.iloc[0]) if pd.notna(row.iloc[0]) else '')
            if not code:
                continue
            
            per_day: dict[date, int] = {}
            for i, cell_value in enumerate(row.iloc[1:]):
                d = date_cols[i]
                if not d:
                    continue
                try:
                    # Handle NaN, None, or empty values
                    if pd.isna(cell_value):
                        q = 0
                    else:
                        q = int(float(cell_value))
                except Exception:
                    q = 0
                per_day[d] = q
            planned[code] = per_day
    except Exception as e:
        print(f"[LOAD] ERROR loading planned XLSX: {e}")
        import traceback
        traceback.print_exc()
        return planned
    
    print(f"[LOAD] Successfully loaded {len(planned)} production lines")
    return planned


def load_planned_from_ltp_csv(path: str) -> dict[str, dict[date, int]]:
    planned: dict[str, dict[date, int]] = {}
    if not path or not os.path.exists(path):
        print(f"[LTP] Fallback CSV not found: {path}")
        return planned

    try:
        df = pd.read_csv(path)
    except Exception as e:
        print(f"[LTP] ERROR reading fallback CSV {path}: {e}")
        return planned

    if df.empty:
        print(f"[LTP] Fallback CSV is empty: {path}")
        return planned

    prod_col = None
    for col in df.columns:
        name = str(col).strip().lower()
        if name in ('prod line', 'prod_line', 'prodline'):
            prod_col = col
            break

    if prod_col is None:
        print("[LTP] Fallback CSV missing 'Prod Line' column")
        return planned

    date_cols: list[tuple[str, date]] = []
    for col in df.columns:
        d = _coerce_header_to_date(col)
        if d:
            date_cols.append((col, d))

    if not date_cols:
        print("[LTP] Fallback CSV has no date columns")
        return planned

    for _, row in df.iterrows():
        code = norm_code(row.get(prod_col))
        if not code:
            continue
        per_day = planned.setdefault(code, {})
        for col, d in date_cols:
            val = row.get(col)
            try:
                if pd.isna(val):
                    qty = 0
                else:
                    qty = int(round(float(val)))
            except Exception:
                qty = 0
            if qty:
                per_day[d] = per_day.get(d, 0) + qty

    print(f"[LTP] Loaded {len(planned)} production lines from fallback CSV")
    return planned


def load_planned_from_ltp_formulas_xlsx(path: str, sheet_name: str = 'LTP_Formulas') -> dict[str, dict[date, int]]:
    planned: dict[str, dict[date, int]] = {}
    if not path or not os.path.exists(path):
        print(f"[LTP-FORM] Workbook not found: {path}")
        return planned

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[LTP-FORM] ERROR opening workbook {path}: {e}")
        return planned

    try:
        resolved_sheet = sheet_name
        if resolved_sheet not in wb.sheetnames:
            resolved_sheet = next((s for s in wb.sheetnames if s.lower() == str(sheet_name).lower()), '')
        if not resolved_sheet or resolved_sheet not in wb.sheetnames:
            print(f"[LTP-FORM] Sheet '{sheet_name}' not found in {os.path.basename(path)}")
            return planned

        ws = wb[resolved_sheet]
        if ws.max_row < 2 or ws.max_column < 2:
            print(f"[LTP-FORM] Sheet '{resolved_sheet}' is empty")
            return planned

        # Expect header row 1: Prod Line, Project, Model, SEW/ASSY, <date columns...>
        header_row = 1
        date_headers: list[tuple[int, date]] = []
        for col in range(1, ws.max_column + 1):
            d = _coerce_header_to_date(ws.cell(row=header_row, column=col).value)
            if d:
                date_headers.append((col, d))
        if not date_headers:
            print(f"[LTP-FORM] No date columns detected in '{resolved_sheet}'")
            return planned

        for r in range(2, ws.max_row + 1):
            code = norm_code(ws.cell(row=r, column=1).value)
            if not code:
                continue
            per_day = planned.setdefault(code, {})
            for col_idx, d in date_headers:
                val = ws.cell(row=r, column=col_idx).value
                try:
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        qty = 0
                    else:
                        qty = int(round(float(val)))
                except Exception:
                    qty = 0
                if qty:
                    per_day[d] = per_day.get(d, 0) + qty
    finally:
        try:
            wb.close()
        except Exception:
            pass

    print(f"[LTP-FORM] Loaded {len(planned)} production lines from formulas workbook")
    return planned


def _ltp_looks_weekly(planned: dict[str, dict[date, int]]) -> bool:
    all_dates: list[date] = []
    for v in planned.values():
        if not v:
            continue
        all_dates.extend(list(v.keys()))

    uniq = sorted(set(all_dates))
    if len(uniq) < 3:
        return False

    monday_ratio = sum(1 for d in uniq if d.weekday() == 0) / float(len(uniq))
    if monday_ratio < 0.7:
        return False

    diffs = [(uniq[i + 1] - uniq[i]).days for i in range(len(uniq) - 1)]
    if not diffs:
        return False
    near_week_ratio = sum(1 for dd in diffs if 6 <= dd <= 8) / float(len(diffs))
    return near_week_ratio >= 0.5


def _expand_weekly_plan_to_daily(
    planned_weekly: dict[str, dict[date, int]],
    workdays_per_week: int,
) -> dict[str, dict[date, int]]:
    if workdays_per_week <= 0:
        return planned_weekly

    expanded: dict[str, dict[date, int]] = {}
    for code, days in planned_weekly.items():
        if not days:
            continue
        out = expanded.setdefault(code, {})
        for week_start, qty in days.items():
            try:
                q = int(qty or 0)
            except Exception:
                q = 0
            if q <= 0:
                continue
            base = q // workdays_per_week
            rem = q % workdays_per_week
            for i in range(workdays_per_week):
                d = week_start + timedelta(days=i)
                add = base + (1 if i < rem else 0)
                if add:
                    out[d] = out.get(d, 0) + add
    return expanded


def _normalize_ltp_text(value: object) -> str:
    s = str(value or '')
    s = s.replace('\xa0', ' ')
    s = s.strip()
    s = re.sub(r'\s+', ' ', s)
    # Normalise common separators so that e.g. "CD / CTE" matches "CD/CTE"
    s = re.sub(r'\s*([/&-])\s*', r'\1', s)
    return s


def _normalize_ltp_key(value: object) -> str:
    return _normalize_ltp_text(value).upper()


def _normalize_ltp_type(value: object) -> str:
    token = _normalize_ltp_key(value)
    if 'ASSY' in token:
        return 'ASSY'
    if 'SEW' in token:
        return 'SEW'
    return ''


def _cell_rgb(cell) -> tuple[int, int, int] | None:
    if cell is None or not cell.fill:
        return None

    color = getattr(cell.fill, 'fgColor', None) or getattr(cell.fill, 'start_color', None)
    if not color:
        return None

    rgb = getattr(color, 'rgb', None)
    if isinstance(rgb, str) and rgb:
        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) != 6:
            return None
        try:
            return (int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16))
        except Exception:
            return None

    idx = getattr(color, 'indexed', None)
    if idx is not None and COLOR_INDEX is not None:
        try:
            idx_int = int(idx)
            if 0 <= idx_int < len(COLOR_INDEX):
                raw = COLOR_INDEX[idx_int]
                if isinstance(raw, str):
                    if len(raw) == 8:
                        raw = raw[2:]
                    if len(raw) == 6:
                        return (int(raw[0:2], 16), int(raw[2:4], 16), int(raw[4:6], 16))
        except Exception:
            return None

    return None


def _infer_type_from_cell(cell) -> str:
    rgb = _cell_rgb(cell)
    if not rgb:
        return ''
    r, g, b = rgb
    if b >= 120 and (b - max(r, g)) >= 40:
        return 'ASSY'
    if r >= 120 and (r - max(g, b)) >= 40:
        return 'SEW'
    return ''


def _infer_type_from_row(ws, row_idx: int, start_col: int, end_col: int) -> str:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        inferred = _infer_type_from_cell(cell)
        if inferred:
            return inferred
    return ''


def _infer_type_from_context(ws, row_idx: int) -> str:
    """Infer SEW/ASSY by scanning nearby header/context cells above the current row.

    Many LTP sheets separate SEW/ASSY into blocks with a header label; styles can be
    conditional-formatting and may not be visible via openpyxl fills.
    """
    scan_up = 40
    scan_cols = 8
    for r in range(row_idx - 1, max(1, row_idx - scan_up) - 1, -1):
        for c in range(1, scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            t = _normalize_ltp_type(v)
            if t:
                return t
    return ''


def _load_ltp_reference(path: str):
    ref_triplet: dict[tuple[str, str, str], list[str]] = {}
    ref_pair: dict[tuple[str, str], list[str]] = {}
    if not path or not os.path.exists(path):
        print(f"[LTP] Reference CSV not found: {path}")
        return ref_triplet, ref_pair

    with open(path, 'r', newline='', encoding='utf-8-sig') as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            code = norm_code(row.get('Prod Line') or row.get('prod_line') or row.get('prod line') or '')
            project = _normalize_ltp_key(
                row.get('Production Line')
                or row.get('production line')
                or row.get('Project')
                or row.get('project')
                or ''
            )
            model = _normalize_ltp_key(row.get('Model') or row.get('model') or '')
            row_type = _normalize_ltp_type(row.get('SEW/ASSY') or row.get('sew/assy') or row.get('type') or '')
            if not code or not project or not model:
                continue
            if row_type:
                key = (project, model, row_type)
                codes = ref_triplet.setdefault(key, [])
                if code not in codes:
                    codes.append(code)
            pair_codes = ref_pair.setdefault((project, model), [])
            if code not in pair_codes:
                pair_codes.append(code)
    return ref_triplet, ref_pair


def _load_ref_meta(path: str) -> dict[str, dict[str, str]]:
    meta: dict[str, dict[str, str]] = {}
    if not path or not os.path.exists(path):
        return meta

    with open(path, 'r', newline='', encoding='utf-8-sig') as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            code = norm_code(row.get('Prod Line') or row.get('prod_line') or row.get('prod line') or '')
            if not code:
                continue

            project_group = (row.get('PROJECT') or row.get('project_group') or row.get('Project Group') or '')
            sew_label = (row.get('SEW') or '')
            assy_label = (row.get('ASSY') or '')
            row_type = _normalize_ltp_type(row.get('SEW/ASSY') or row.get('sew/assy') or row.get('type') or '')

            m = meta.setdefault(code, {})
            if project_group and not m.get('project_group'):
                m['project_group'] = str(project_group).strip()
            if row_type and not m.get('type'):
                m['type'] = row_type
            if sew_label and not m.get('sew'):
                m['sew'] = str(sew_label).strip()
            if assy_label and not m.get('assy'):
                m['assy'] = str(assy_label).strip()
    return meta


def _load_ltp_page_reference(path: str) -> dict[tuple[str, str, str], dict[str, str | float]]:
    ref: dict[tuple[str, str, str], dict[str, str | float]] = {}
    if not path or not os.path.exists(path):
        print(f"[LTP] Reference CSV not found: {path}")
        return ref

    with open(path, 'r', newline='', encoding='utf-8-sig') as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            project = _normalize_ltp_key(
                row.get('Production Line')
                or row.get('production line')
                or row.get('Project')
                or row.get('project')
                or ''
            )
            model = _normalize_ltp_key(row.get('Model') or row.get('model') or '')
            row_type = _normalize_ltp_type(row.get('SEW/ASSY') or row.get('sew/assy') or row.get('type') or '')
            if not project or not model or not row_type:
                continue

            # Parse LTP multiplier (default 1.0 if not specified or invalid)
            multiplier_raw = row.get('LTP multiplier') or row.get('ltp multiplier') or ''
            try:
                multiplier = float(multiplier_raw) if multiplier_raw else 1.0
                if multiplier <= 0:
                    multiplier = 1.0
            except (ValueError, TypeError):
                multiplier = 1.0

            ref[(project, model, row_type)] = {
                'PROJECT': str(row.get('PROJECT') or '').strip(),
                'SEW': str(row.get('SEW') or '').strip(),
                'ASSY': str(row.get('ASSY') or '').strip(),
                'multiplier': multiplier,
            }

    return ref


def load_planned_pages_from_ltp(
    directory: str,
    sheet_name: str,
    label: str,
    ref_csv: str,
    keywords,
    date_row: int,
    date_start_col: str,
    date_end_col: str,
    workdays_per_week: int,
    fallback_csv: str | None = None,
) -> dict[str, dict[str, dict[date, int]]]:
    """Return planned quantities aggregated to page labels:
    { 'PROJECT': {label: {date: qty}}, 'SEW': {...}, 'ASSY': {...} }
    """
    result: dict[str, dict[str, dict[date, int]]] = {'PROJECT': {}, 'SEW': {}, 'ASSY': {}}

    workbook_path = _find_ltp_workbook(directory, keywords)
    if not workbook_path:
        print("[LTP] No workbook found for planned page exports; skipping")
        return result

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)
    except Exception as e:
        print(f"[LTP] ERROR opening workbook {workbook_path}: {e}")
        return result

    try:
        desired = (sheet_name or '').strip()
        candidates: list[str] = []
        if desired:
            candidates.append(desired)
        candidates.extend(['Planned', 'Planning', 'PLANING'])
        resolved_sheet = ''
        for cand in candidates:
            if cand in wb.sheetnames:
                resolved_sheet = cand
                break
            lower_match = next((s for s in wb.sheetnames if s.lower() == cand.lower()), '')
            if lower_match:
                resolved_sheet = lower_match
                break

        if not resolved_sheet:
            print(f"[LTP] Sheet '{sheet_name}' not found in {workbook_path.name} (available: {wb.sheetnames})")
            return result

        ws = wb[resolved_sheet]
        start_col_idx = column_index_from_string(str(date_start_col).strip() or 'X')
        end_col_idx = column_index_from_string(str(date_end_col).strip() or 'BW')
        if end_col_idx < start_col_idx:
            start_col_idx, end_col_idx = end_col_idx, start_col_idx

        date_headers: list[date | None] = []
        for col in range(start_col_idx, end_col_idx + 1):
            date_headers.append(_coerce_header_to_date(ws.cell(row=date_row, column=col).value))

        if sum(1 for d in date_headers if d) < 3:
            scan_max = min(ws.max_row, 200)
            best_row = None
            best_count = 0
            for r in range(1, scan_max + 1):
                cnt = 0
                for col in range(start_col_idx, end_col_idx + 1):
                    if _coerce_header_to_date(ws.cell(row=r, column=col).value):
                        cnt += 1
                if cnt > best_count:
                    best_count = cnt
                    best_row = r

            if best_row is not None and best_count >= 3:
                print(f"[LTP] Date header row auto-detected: {best_row} ({best_count} date-like columns)")
                date_row = best_row
                date_headers = []
                for col in range(start_col_idx, end_col_idx + 1):
                    date_headers.append(_coerce_header_to_date(ws.cell(row=date_row, column=col).value))

        ref_pages = _load_ltp_page_reference(ref_csv)
        target_label = (label or '').strip().lower()

        label_col_idx: int | None = 4 if target_label else None
        if target_label:
            scan_rows = min(ws.max_row, 500)
            direct_hits = 0
            for r in range(1, scan_rows + 1):
                v = ws.cell(row=r, column=4).value
                if str(v or '').strip().lower() == target_label:
                    direct_hits += 1
            if direct_hits == 0:
                best_col = None
                best_hits = 0
                scan_cols = min(ws.max_column, 15)
                for c in range(1, scan_cols + 1):
                    hits = 0
                    for r in range(1, scan_rows + 1):
                        v = ws.cell(row=r, column=c).value
                        if str(v or '').strip().lower() == target_label:
                            hits += 1
                    if hits > best_hits:
                        best_hits = hits
                        best_col = c
                if best_col is not None and best_hits > 0:
                    label_col_idx = best_col
                    print(f"[LTP] Label column auto-detected: {label_col_idx} ({best_hits} matches)")
                else:
                    label_col_idx = None
                    print(f"[LTP] ERROR: Label '{label}' not found in sheet '{resolved_sheet}'; planned page exports skipped")
                    return result

        for row_idx in range(1, ws.max_row + 1):
            if label_col_idx is not None:
                label_raw = ws.cell(row=row_idx, column=label_col_idx).value
                if (str(label_raw or '').strip().lower()) != target_label:
                    continue

            project_raw = ws.cell(row=row_idx, column=1).value
            model_raw = ws.cell(row=row_idx, column=2).value
            project_key = _normalize_ltp_key(project_raw)
            model_key = _normalize_ltp_key(model_raw)
            if not project_key or not model_key:
                continue

            row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=3).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=5).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=6).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=4).value)
            if not row_type:
                row_type = _normalize_ltp_type(project_key) or _normalize_ltp_type(model_key)
            if not row_type:
                row_type = _infer_type_from_row(ws, row_idx, start_col_idx, end_col_idx)
            if not row_type:
                row_type = _infer_type_from_row(ws, row_idx, 1, max(end_col_idx, 10))
            if not row_type:
                row_type = _infer_type_from_context(ws, row_idx)
            if not row_type:
                continue

            labels = ref_pages.get((project_key, model_key, row_type))
            if not labels:
                other = 'ASSY' if row_type == 'SEW' else 'SEW'
                labels = ref_pages.get((project_key, model_key, other))
                if labels:
                    print(f"[LTP] INFO: Using {other} label mapping for {project_key}/{model_key} (inferred {row_type})")
                else:
                    print(f"[LTP] WARNING: No label mapping for {project_key}/{model_key} ({row_type})")
                    continue

            per_day: dict[date, int] = {}
            for offset, col in enumerate(range(start_col_idx, end_col_idx + 1)):
                d = date_headers[offset] if offset < len(date_headers) else None
                if not d:
                    continue
                val = ws.cell(row=row_idx, column=col).value
                try:
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        qty = 0
                    else:
                        qty = int(round(float(val)))
                except Exception:
                    qty = 0
                if qty:
                    per_day[d] = per_day.get(d, 0) + qty

            if not per_day:
                continue

            # Apply LTP multiplier from ref.csv (CV=2x, PZ1D=7x, default=1x)
            multiplier = float(labels.get('multiplier', 1.0) or 1.0)
            if multiplier != 1.0:
                per_day = {d: int(round(qty * multiplier)) for d, qty in per_day.items()}
                print(f"[LTP] Applied multiplier {multiplier}x for {project_key}/{model_key} ({row_type})")

            proj_label = (labels.get('PROJECT') or project_key).strip()
            if proj_label:
                bucket = result['PROJECT'].setdefault(proj_label, {})
                for d, qty in per_day.items():
                    bucket[d] = bucket.get(d, 0) + qty

            sew_label = (labels.get('SEW') or '').strip()
            if sew_label:
                bucket = result['SEW'].setdefault(sew_label, {})
                for d, qty in per_day.items():
                    bucket[d] = bucket.get(d, 0) + qty

            assy_label = (labels.get('ASSY') or '').strip()
            if assy_label:
                bucket = result['ASSY'].setdefault(assy_label, {})
                for d, qty in per_day.items():
                    bucket[d] = bucket.get(d, 0) + qty

        for key in ('PROJECT', 'SEW', 'ASSY'):
            series = result.get(key) or {}
            if series and _ltp_looks_weekly(series):
                result[key] = _expand_weekly_plan_to_daily(series, workdays_per_week)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return result


def _write_monthly_csv_by_label(out_path: str, series_by_label: dict[str, dict[date, int]], month_start: date):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    days_in_month = calendar.monthrange(month_start.year, month_start.month)[1]
    dates = [month_start.replace(day=i) for i in range(1, days_in_month + 1)]

    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        f.write('Label,' + ','.join(d.isoformat() for d in dates) + "\n")
        for label in sorted(series_by_label.keys(), key=lambda s: str(s).upper()):
            per_day = series_by_label.get(label, {})
            row = [str(label)]
            for d in dates:
                row.append(f"{float(per_day.get(d, 0) or 0):.2f}")
            f.write(','.join(row) + "\n")
        f.write("\n")


def _write_weekly_csv_by_label(out_path: str, series_by_label: dict[str, dict[date, int]], month_start: date):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    days_in_month = calendar.monthrange(month_start.year, month_start.month)[1]
    month_end = month_start.replace(day=days_in_month)
    week_starts: list[date] = []
    seen: set[date] = set()
    for d in daterange(month_start, month_end):
        ws = monday_of_week(d)
        if ws not in seen:
            seen.add(ws)
            week_starts.append(ws)
    week_starts.sort()

    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        f.write('Label,' + ','.join(d.isoformat() for d in week_starts) + "\n")
        for label in sorted(series_by_label.keys(), key=lambda s: str(s).upper()):
            per_day = series_by_label.get(label, {})
            row = [str(label)]
            for ws in week_starts:
                we = ws + timedelta(days=6)
                row.append(f"{float(aggregate(per_day, ws, we) or 0):.2f}")
            f.write(','.join(row) + "\n")
        f.write("\n")


def _write_month_total_csv_by_label(out_path: str, series_by_label: dict[str, dict[date, int]], month_start: date):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    days_in_month = calendar.monthrange(month_start.year, month_start.month)[1]
    month_end = month_start.replace(day=days_in_month)
    month_key = f"{month_start.year:04d}-{month_start.month:02d}"

    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        f.write('Label,' + month_key + "\n")
        for label in sorted(series_by_label.keys(), key=lambda s: str(s).upper()):
            per_day = series_by_label.get(label, {})
            total = float(aggregate(per_day, month_start, month_end) or 0)
            f.write(f"{label},{total:.2f}\n")
        f.write("\n")


def _load_monthly_per_day_csv(path: str, month_start: date) -> dict[str, dict[date, float]]:
    """Parse monthly per-day CSV format: Label,Day1,Day2,... into {label: {date: qty}}."""
    series: dict[str, dict[date, float]] = {}
    if not path or not os.path.exists(path):
        return series

    days_in_month = calendar.monthrange(month_start.year, month_start.month)[1]
    dates = [month_start.replace(day=i) for i in range(1, days_in_month + 1)]

    with open(path, 'r', newline='', encoding='utf-8-sig') as f:
        for raw in f:
            line = (raw or '').strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(',')]
            if not parts or not parts[0]:
                continue

            if _norm_key(parts[0]) in ('LABEL', 'LINE'):
                continue
            if len(parts) > 1 and re.match(r'^\d{4}-\d{2}-\d{2}$', parts[1] or ''):
                continue

            label = str(parts[0]).lstrip('\ufeff').strip()
            per_day = series.setdefault(label, {})
            for idx, d in enumerate(dates, start=1):
                if idx >= len(parts):
                    break
                try:
                    v = float(parts[idx] or 0)
                except Exception:
                    v = 0.0
                if v:
                    per_day[d] = per_day.get(d, 0.0) + v

    return series


def _norm_key(s: str) -> str:
    s2 = (s or '')
    s2 = s2.replace('\ufeff', '')
    s2 = s2.replace('\xa0', ' ')
    s2 = s2.strip().upper()
    s2 = ' '.join(s2.split())
    s2 = re.sub(r'\s*-\s*', ' - ', s2)
    s2 = ' '.join(s2.split())
    return s2


def _base_label_name(name: str) -> str:
    s = _norm_key(name)
    for suf in (' - SEW', ' - ASSY'):
        if s.endswith(suf):
            s = s[:-len(suf)].strip()
            break
    if ' - ' in s:
        s = s.split(' - ')[0].strip()
    while s.endswith('-'):
        s = s[:-1].strip()
    return s


def _build_olk_norm_by_label(mapping: dict[str, str], raw_olk: dict[str, float] | None) -> dict[str, float]:
    out: dict[str, float] = {}
    if not raw_olk:
        return out

    code_to_disp: dict[str, str] = {}
    disp_norm_to_disp: dict[str, str] = {}
    for code, disp in (mapping or {}).items():
        if not code or not disp:
            continue
        code_to_disp[_norm_key(code)] = disp
        disp_norm_to_disp[_norm_key(disp)] = disp

    code_keys = list(code_to_disp.keys())
    disp_keys = list(disp_norm_to_disp.keys())

    for key, val in raw_olk.items():
        if not val:
            continue
        key_norm = _norm_key(str(key))
        target_disp: str | None = None
        if key_norm in code_to_disp:
            target_disp = code_to_disp[key_norm]
        elif key_norm in disp_norm_to_disp:
            target_disp = disp_norm_to_disp[key_norm]
        else:
            for ck in code_keys:
                if ck and ck in key_norm:
                    target_disp = code_to_disp.get(ck)
                    break
        if target_disp is None:
            for dk in disp_keys:
                if dk and dk in key_norm:
                    target_disp = disp_norm_to_disp.get(dk)
                    break

        out_key = _norm_key(target_disp) if target_disp else key_norm
        out[out_key] = out.get(out_key, 0.0) + float(val or 0.0)

    return out


def _olk_lookup(label: str, olk_norm: dict[str, float]) -> float:
    if not label or not olk_norm:
        return 0.0
    k = _norm_key(label)
    if k in olk_norm:
        return float(olk_norm.get(k, 0.0) or 0.0)
    base = _base_label_name(label)
    if base in olk_norm:
        return float(olk_norm.get(base, 0.0) or 0.0)
    return 0.0


def _load_master_list(path: str) -> dict[str, list[str]]:
    res: dict[str, list[str]] = {'PROJECT': [], 'SEW': [], 'ASSY': []}
    if not path or not os.path.exists(path):
        return res
    with open(path, 'r', newline='', encoding='utf-8-sig') as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            page = str(row.get('Dashboard page name') or row.get('page') or '').strip().upper()
            label = str(row.get('Rows') or row.get('row') or row.get('label') or '').strip()
            if not page or not label:
                continue
            if page in res:
                res[page].append(label)
    return res


def _canonicalize_series(
    series: dict[str, dict[date, float]],
    canonical: list[str],
) -> dict[str, dict[date, float]]:
    if not canonical:
        return series
    canon_by_norm = {_norm_key(c): c for c in canonical}
    out: dict[str, dict[date, float]] = {}
    for label, per_day in (series or {}).items():
        key_norm = _norm_key(str(label))
        target = canon_by_norm.get(key_norm) or str(label)
        tgt = out.setdefault(target, {})
        for d, v in (per_day or {}).items():
            tgt[d] = float(tgt.get(d, 0.0) or 0.0) + float(v or 0.0)
    for c in canonical:
        out.setdefault(c, {})
    return out


def _compute_metrics_from_page_csvs(
    as_of: date,
    daily_start: date,
    start_week: date,
    start_month: date,
) -> dict[str, object] | None:
    prod_dir = os.path.join(_BASE_DIR, 'PVS', 'Production')
    plan_dir = os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Day')
    if not os.path.exists(os.path.join(plan_dir, '1_PVS_per_Project.csv')):
        plan_dir = os.path.join(_BASE_DIR, 'PVS', 'Planned')

    planned_project_raw = _load_monthly_per_day_csv(os.path.join(plan_dir, '1_PVS_per_Project.csv'), start_month)
    planned_sew_raw = _load_monthly_per_day_csv(os.path.join(plan_dir, '2_PVS_per_SEW.csv'), start_month)
    planned_assy_raw = _load_monthly_per_day_csv(os.path.join(plan_dir, '3_PVS_per_ASSY.csv'), start_month)

    produced_project_raw = _load_monthly_per_day_csv(os.path.join(prod_dir, '1_PVS_per_Project.csv'), start_month)
    produced_sew_raw = _load_monthly_per_day_csv(os.path.join(prod_dir, '2_PVS_per_SEW.csv'), start_month)
    produced_assy_raw = _load_monthly_per_day_csv(os.path.join(prod_dir, '3_PVS_per_ASSY.csv'), start_month)

    if not (planned_project_raw or planned_sew_raw or planned_assy_raw or produced_project_raw or produced_sew_raw or produced_assy_raw):
        return None

    master = _load_master_list(os.path.join(_BASE_DIR, 'PVS', 'master_list.csv'))
    canonical_project = master.get('PROJECT') or []
    canonical_sew = master.get('SEW') or []
    canonical_assy = master.get('ASSY') or []

    canon_proj_norm = {_norm_key(x) for x in canonical_project}
    canon_sew_norm = {_norm_key(x) for x in canonical_sew}
    canon_assy_norm = {_norm_key(x) for x in canonical_assy}

    plan_proj_norm = {_norm_key(k) for k in planned_project_raw.keys()}
    plan_sew_norm = {_norm_key(k) for k in planned_sew_raw.keys()}
    plan_assy_norm = {_norm_key(k) for k in planned_assy_raw.keys()}
    prod_proj_norm = {_norm_key(k) for k in produced_project_raw.keys()}
    prod_sew_norm = {_norm_key(k) for k in produced_sew_raw.keys()}
    prod_assy_norm = {_norm_key(k) for k in produced_assy_raw.keys()}

    if canonical_project:
        missing = [c for c in canonical_project if _norm_key(c) not in plan_proj_norm]
        if missing:
            print(f"[CSV] Missing planned PROJECT rows: {missing}")
        missing = [c for c in canonical_project if _norm_key(c) not in prod_proj_norm]
        if missing:
            print(f"[CSV] Missing production PROJECT rows: {missing}")
        extra = [k for k in planned_project_raw.keys() if _norm_key(k) not in canon_proj_norm]
        if extra:
            print(f"[CSV] Extra planned PROJECT rows (not in master): {extra}")
        extra = [k for k in produced_project_raw.keys() if _norm_key(k) not in canon_proj_norm]
        if extra:
            print(f"[CSV] Extra production PROJECT rows (not in master): {extra}")

    if canonical_sew:
        missing = [c for c in canonical_sew if _norm_key(c) not in plan_sew_norm]
        if missing:
            print(f"[CSV] Missing planned SEW rows: {missing}")
        missing = [c for c in canonical_sew if _norm_key(c) not in prod_sew_norm]
        if missing:
            print(f"[CSV] Missing production SEW rows: {missing}")
        extra = [k for k in planned_sew_raw.keys() if _norm_key(k) not in canon_sew_norm]
        if extra:
            print(f"[CSV] Extra planned SEW rows (not in master): {extra}")
        extra = [k for k in produced_sew_raw.keys() if _norm_key(k) not in canon_sew_norm]
        if extra:
            print(f"[CSV] Extra production SEW rows (not in master): {extra}")

    if canonical_assy:
        missing = [c for c in canonical_assy if _norm_key(c) not in plan_assy_norm]
        if missing:
            print(f"[CSV] Missing planned ASSY rows: {missing}")
        missing = [c for c in canonical_assy if _norm_key(c) not in prod_assy_norm]
        if missing:
            print(f"[CSV] Missing production ASSY rows: {missing}")
        extra = [k for k in planned_assy_raw.keys() if _norm_key(k) not in canon_assy_norm]
        if extra:
            print(f"[CSV] Extra planned ASSY rows (not in master): {extra}")
        extra = [k for k in produced_assy_raw.keys() if _norm_key(k) not in canon_assy_norm]
        if extra:
            print(f"[CSV] Extra production ASSY rows (not in master): {extra}")

    planned_project = _canonicalize_series(planned_project_raw, canonical_project)
    planned_sew = _canonicalize_series(planned_sew_raw, canonical_sew)
    planned_assy = _canonicalize_series(planned_assy_raw, canonical_assy)
    produced_project = _canonicalize_series(produced_project_raw, canonical_project)
    produced_sew = _canonicalize_series(produced_sew_raw, canonical_sew)
    produced_assy = _canonicalize_series(produced_assy_raw, canonical_assy)

    mapping = load_map_csv(PVS_MAP_CSV)
    olk_norm = _build_olk_norm_by_label(mapping, load_olk_csv(PVS_OLK_CSV))

    def adherence(delta: float, schedule: float) -> float | None:
        if schedule <= 0:
            return None
        pct = (delta / schedule) * 100.0
        if pct > PVS_ADHERENCE_CLAMP:
            return PVS_ADHERENCE_CLAMP
        if pct < -PVS_ADHERENCE_CLAMP:
            return -PVS_ADHERENCE_CLAMP
        return pct

    rows: list[dict[str, object]] = []

    label_order_sew = {_norm_key(v): i for i, v in enumerate(canonical_sew)} if canonical_sew else {}
    label_order_assy = {_norm_key(v): i for i, v in enumerate(canonical_assy)} if canonical_assy else {}

    def _sorted_labels(category: str, labels: set[str]) -> list[str]:
        if category == 'SEW' and label_order_sew:
            return sorted(labels, key=lambda s: (label_order_sew.get(_norm_key(str(s)), 10**9), _norm_key(str(s))))
        if category == 'ASSY' and label_order_assy:
            return sorted(labels, key=lambda s: (label_order_assy.get(_norm_key(str(s)), 10**9), _norm_key(str(s))))
        return sorted(labels, key=lambda s: _norm_key(str(s)))

    def _append_rows(category: str, planned: dict[str, dict[date, float]], produced: dict[str, dict[date, float]]):
        labels = _sorted_labels(category, set(planned.keys()) | set(produced.keys()))
        for label in labels:
            plan_days = planned.get(label, {})
            prod_days = produced.get(label, {})

            plan_day = int(aggregate(plan_days, daily_start, as_of))
            plan_wtd = int(aggregate(plan_days, start_week, as_of))
            plan_mtd = int(aggregate(plan_days, start_month, as_of))

            prod_day = float(aggregate(prod_days, daily_start, as_of))
            prod_wtd = float(aggregate(prod_days, start_week, as_of))
            prod_mtd = float(aggregate(prod_days, start_month, as_of))

            d_day = prod_day - plan_day
            d_wtd = prod_wtd - plan_wtd
            d_mtd = prod_mtd - plan_mtd

            line_olk = float(_olk_lookup(str(label), olk_norm) or 0.0)
            if line_olk > 0:
                adh_olk_pct = (prod_mtd / line_olk) * 100.0
            else:
                adh_olk_pct = 0.0

            adh_mtd = adherence(d_mtd, plan_mtd)
            adh_wtd = adherence(d_wtd, plan_wtd)
            adh_day = adherence(d_day, plan_day)

            rows.append({
                'code': str(label),
                'line': str(label),
                'category': category,
                'mtd': {
                    'olk': int(round(line_olk)),
                    'adh_olk_pct': round(adh_olk_pct, 1),
                    'schedule': plan_mtd,
                    'production': round(prod_mtd, 2),
                    'delta': round(d_mtd, 2),
                    'adherence_pct': round(adh_mtd, 1) if adh_mtd is not None else None,
                },
                'wtd': {
                    'schedule': plan_wtd,
                    'production': round(prod_wtd, 2),
                    'delta': round(d_wtd, 2),
                    'adherence_pct': round(adh_wtd, 1) if adh_wtd is not None else None,
                },
                'daily': {
                    'schedule': plan_day,
                    'production': round(prod_day, 2),
                    'delta': round(d_day, 2),
                    'adherence_pct': round(adh_day, 1) if adh_day is not None else None,
                },
            })

    _append_rows('SEW', planned_sew, produced_sew)
    _append_rows('ASSY', planned_assy, produced_assy)

    category_order = {'SEW': 0, 'ASSY': 1, 'OTHER': 2}
    def _row_sort_key(r: dict[str, object]):
        cat = str(r.get('category', 'OTHER'))
        line = str(r.get('line', ''))
        if cat == 'SEW' and label_order_sew:
            return (category_order.get(cat, 2), label_order_sew.get(_norm_key(line), 10**9), line)
        if cat == 'ASSY' and label_order_assy:
            return (category_order.get(cat, 2), label_order_assy.get(_norm_key(line), 10**9), line)
        return (category_order.get(cat, 2), line)
    rows.sort(key=_row_sort_key)

    expected_anoms: list[str] = []
    expected_sew_norm = {_norm_key(x) for x in canonical_sew}
    expected_assy_norm = {_norm_key(x) for x in canonical_assy}
    for r in rows:
        cat = str(r.get('category', ''))
        line = str(r.get('line', ''))
        if cat == 'SEW' and expected_sew_norm and _norm_key(line) not in expected_sew_norm:
            continue
        if cat == 'ASSY' and expected_assy_norm and _norm_key(line) not in expected_assy_norm:
            continue
        for win_key in ('daily', 'wtd', 'mtd'):
            win = r.get(win_key) or {}
            if not isinstance(win, dict):
                continue
            sched = float(win.get('schedule', 0) or 0)
            prod = float(win.get('production', 0) or 0)
            if sched <= 0 and prod > 0:
                expected_anoms.append(f"{cat} {line} {win_key}: schedule=0 production={prod}")
            elif sched > 0 and prod <= 0:
                expected_anoms.append(f"{cat} {line} {win_key}: schedule={sched} production=0")
    if expected_anoms:
        print('[CSV] Zero anomalies (expected rows):')
        for s in expected_anoms:
            print(f"[CSV] {s}")

    def _empty_bucket():
        return {'schedule': 0, 'production': 0}

    totals = {
        'sew': {'mtd': _empty_bucket(), 'wtd': _empty_bucket(), 'daily': _empty_bucket()},
        'assy': {'mtd': _empty_bucket(), 'wtd': _empty_bucket(), 'daily': _empty_bucket()},
        'other': {'mtd': _empty_bucket(), 'wtd': _empty_bucket(), 'daily': _empty_bucket()},
        'all': {'mtd': _empty_bucket(), 'wtd': _empty_bucket(), 'daily': _empty_bucket()},
    }

    olk_totals = {
        'sew': {'olk': 0.0, 'production': 0.0},
        'assy': {'olk': 0.0, 'production': 0.0},
    }

    def _accumulate(bucket: dict, key: str, sched: int | float, prod: int | float) -> None:
        b = bucket[key]
        b['schedule'] += int(sched or 0)
        b['production'] += float(prod or 0.0)

    for r in rows:
        cat = str(r.get('category', 'OTHER'))
        _accumulate(totals['all'], 'mtd', r['mtd']['schedule'], r['mtd']['production'])
        _accumulate(totals['all'], 'wtd', r['wtd']['schedule'], r['wtd']['production'])
        _accumulate(totals['all'], 'daily', r['daily']['schedule'], r['daily']['production'])

        if cat == 'SEW':
            bucket = totals['sew']
            olk_key = 'sew'
        elif cat == 'ASSY':
            bucket = totals['assy']
            olk_key = 'assy'
        else:
            bucket = totals['other']
            olk_key = None

        _accumulate(bucket, 'mtd', r['mtd']['schedule'], r['mtd']['production'])
        _accumulate(bucket, 'wtd', r['wtd']['schedule'], r['wtd']['production'])
        _accumulate(bucket, 'daily', r['daily']['schedule'], r['daily']['production'])

        if olk_key is not None:
            line_olk = float(r['mtd'].get('olk', 0) or 0.0)
            if line_olk:
                cat_olk = olk_totals[olk_key]
                cat_olk['olk'] += line_olk
                cat_olk['production'] += float(r['mtd'].get('production', 0.0) or 0.0)

    base_olk_by_project: dict[str, float] = {}
    for r in rows:
        base = _base_label_name(str(r.get('line', '')))
        if not base:
            continue
        v = float(r.get('mtd', {}).get('olk', 0.0) or 0.0)  # type: ignore[union-attr]
        if v > float(base_olk_by_project.get(base, 0.0) or 0.0):
            base_olk_by_project[base] = v

    group_buckets: dict[str, dict[str, dict[str, object]]] = {}
    if canonical_project:
        project_labels = list(canonical_project)
    else:
        project_labels = sorted(set(planned_project.keys()) | set(produced_project.keys()), key=lambda s: _norm_key(str(s)))
    for label in project_labels:
        base = _base_label_name(str(label))
        if not base:
            continue
        group_name = GROUP_OVERRIDES.get(base) or base

        plan_days = planned_project.get(label, {})
        prod_days = produced_project.get(label, {})

        plan_day = float(aggregate(plan_days, daily_start, as_of))
        plan_wtd = float(aggregate(plan_days, start_week, as_of))
        plan_mtd = float(aggregate(plan_days, start_month, as_of))

        prod_day = float(aggregate(prod_days, daily_start, as_of))
        prod_wtd = float(aggregate(prod_days, start_week, as_of))
        prod_mtd = float(aggregate(prod_days, start_month, as_of))

        grp = group_buckets.setdefault(group_name, {
            'mtd': {'schedule': 0.0, 'production': 0.0, 'olk_projects': {}},
            'wtd': {'schedule': 0.0, 'production': 0.0},
            'daily': {'schedule': 0.0, 'production': 0.0},
        })

        grp['daily']['schedule'] = float(grp['daily']['schedule']) + plan_day
        grp['daily']['production'] = float(grp['daily']['production']) + prod_day
        grp['wtd']['schedule'] = float(grp['wtd']['schedule']) + plan_wtd
        grp['wtd']['production'] = float(grp['wtd']['production']) + prod_wtd
        grp['mtd']['schedule'] = float(grp['mtd']['schedule']) + plan_mtd
        grp['mtd']['production'] = float(grp['mtd']['production']) + prod_mtd

        olk_val = float(base_olk_by_project.get(base, 0.0) or 0.0)
        if olk_val:
            olk_projects = grp['mtd'].setdefault('olk_projects', {})
            if isinstance(olk_projects, dict):
                prev = float(olk_projects.get(base, 0.0) or 0.0)
                if olk_val > prev:
                    olk_projects[base] = olk_val

    group_totals: list[dict[str, object]] = []
    for name in sorted(group_buckets.keys()):
        src = group_buckets[name]
        agg: dict[str, object] = {'group': name}
        for key in ('mtd', 'wtd', 'daily'):
            win = src[key]
            sched = float(win.get('schedule', 0.0) or 0.0)
            prod = float(win.get('production', 0.0) or 0.0)
            delta = prod - sched
            adh = adherence(delta, sched)
            win_out: dict[str, object] = {
                'schedule': int(round(sched)),
                'production': round(prod, 2),
                'delta': round(delta, 2),
                'adherence_pct': round(adh, 1) if adh is not None else None,
            }
            if key == 'mtd':
                olk_projects = win.get('olk_projects')
                if isinstance(olk_projects, dict) and olk_projects:
                    olk_qty = sum(float(v or 0.0) for v in olk_projects.values())
                else:
                    olk_qty = 0.0
                if olk_qty > 0:
                    adh_olk_pct = (prod / olk_qty) * 100.0
                else:
                    adh_olk_pct = 0.0
                win_out['olk'] = int(round(olk_qty))
                win_out['adh_olk_pct'] = round(adh_olk_pct, 1)
            agg[key] = win_out
        group_totals.append(agg)

    olk_totals_out: dict[str, dict[str, float]] = {}
    for key in ('sew', 'assy'):
        src = olk_totals.get(key, {})
        olk_qty = float(src.get('olk', 0.0) or 0.0)
        prod_qty = float(src.get('production', 0.0) or 0.0)
        if olk_qty > 0:
            adh_olk_pct = (prod_qty / olk_qty) * 100.0
        else:
            adh_olk_pct = 0.0
        olk_totals_out[key] = {
            'olk': int(round(olk_qty)),
            'production': round(prod_qty, 2),
            'adh_olk_pct': round(adh_olk_pct, 1),
        }

    # --- Diagnostic CSV export ---------------------------------------------------
    try:
        diag_dir = os.path.join(_BASE_DIR, 'PVS', 'Debug')
        os.makedirs(diag_dir, exist_ok=True)
        diag_path = os.path.join(diag_dir, 'dashboard_diagnostic.csv')
        with open(diag_path, 'w', newline='', encoding='utf-8') as df:
            df.write('Page,Label,Daily_Sched,Daily_Prod,Daily_Delta,WTD_Sched,WTD_Prod,WTD_Delta,MTD_Sched,MTD_Prod,MTD_Delta,MTD_Adh%,OLK,OLK_Adh%\n')
            # Page 1 – PROJECT group totals
            for gt in group_totals:
                name = gt.get('group', '')
                mtd = gt.get('mtd', {})
                wtd = gt.get('wtd', {})
                daily = gt.get('daily', {})
                df.write(','.join([
                    'PROJECT', str(name),
                    str(daily.get('schedule', 0)), f"{daily.get('production', 0):.2f}", f"{daily.get('delta', 0):.2f}",
                    str(wtd.get('schedule', 0)), f"{wtd.get('production', 0):.2f}", f"{wtd.get('delta', 0):.2f}",
                    str(mtd.get('schedule', 0)), f"{mtd.get('production', 0):.2f}", f"{mtd.get('delta', 0):.2f}",
                    str(mtd.get('adherence_pct', '') if mtd.get('adherence_pct') is not None else ''),
                    str(mtd.get('olk', 0)), str(mtd.get('adh_olk_pct', 0)),
                ]) + '\n')
            # Pages 2 & 3 – SEW / ASSY detail rows
            for r in rows:
                cat = str(r.get('category', ''))
                line = str(r.get('line', ''))
                mtd = r.get('mtd', {})
                wtd = r.get('wtd', {})
                daily = r.get('daily', {})
                df.write(','.join([
                    cat, line,
                    str(daily.get('schedule', 0)), f"{daily.get('production', 0):.2f}", f"{daily.get('delta', 0):.2f}",
                    str(wtd.get('schedule', 0)), f"{wtd.get('production', 0):.2f}", f"{wtd.get('delta', 0):.2f}",
                    str(mtd.get('schedule', 0)), f"{mtd.get('production', 0):.2f}", f"{mtd.get('delta', 0):.2f}",
                    str(mtd.get('adherence_pct', '') if mtd.get('adherence_pct') is not None else ''),
                    str(mtd.get('olk', 0)), str(mtd.get('adh_olk_pct', 0)),
                ]) + '\n')
        print(f"[DIAG] Wrote diagnostic CSV: {diag_path}")
    except Exception as e:
        print(f"[DIAG] WARNING: Could not write diagnostic CSV: {e}")

    return {
        'success': True,
        'date': as_of.strftime('%Y-%m-%d'),
        'rows': rows,
        'totals': totals,
        'group_totals': group_totals,
        'olk_totals': olk_totals_out,
    }


def _find_ltp_workbook(directory: str, keywords) -> Path | None:
    if not directory:
        return None
    dir_path = Path(directory)
    if not dir_path.exists():
        print(f"[LTP] Directory not found: {directory}")
        return None

    excel_exts = {'.xlsx', '.xlsm'}
    files = [
        p
        for p in dir_path.iterdir()
        if p.is_file()
        and p.suffix.lower() in excel_exts
        and not p.name.startswith('~$')
    ]
    if not files:
        print(f"[LTP] No Excel files found in: {directory}")
        return None

    if isinstance(keywords, str):
        keywords = [keywords]
    kw_norm = [str(k).strip().upper() for k in (keywords or []) if str(k).strip()]

    if kw_norm:
        # Use scoring rather than strict all-keywords match to tolerate filename variations
        # (e.g. missing CW token, double spaces, hyphens, etc.).
        scored: list[tuple[int, Path]] = []
        for p in files:
            name_u = p.name.upper()
            score = sum(1 for k in kw_norm if k in name_u)
            if score > 0:
                scored.append((score, p))

        if scored:
            max_score = max(s for s, _ in scored)
            best = [p for s, p in scored if s == max_score]
            if len(best) == 1:
                return best[0]
            latest = max(best, key=lambda p: p.stat().st_mtime)
            print(
                f"[LTP] Multiple keyword matches found (score={max_score}/{len(kw_norm)}); using latest: {latest.name}"
            )
            return latest

    if len(files) == 1:
        return files[0]

    print(f"[LTP] No unique workbook found in {directory} using keywords {kw_norm}")
    return None


def load_planned_from_ltp(
    directory: str,
    sheet_name: str,
    label: str,
    ref_csv: str,
    keywords,
    date_row: int,
    date_start_col: str,
    date_end_col: str,
    fallback_csv: str | None = None,
):
    planned: dict[str, dict[date, int]] = {}
    workbook_path = _find_ltp_workbook(directory, keywords)
    if not workbook_path:
        print("[LTP] No workbook found; falling back to CSV")
        return load_planned_from_ltp_csv(fallback_csv) if fallback_csv else planned

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True)
    except Exception as e:
        print(f"[LTP] ERROR opening workbook {workbook_path}: {e}")
        print("[LTP] Falling back to CSV")
        return load_planned_from_ltp_csv(fallback_csv) if fallback_csv else planned

    try:
        # Resolve sheet name with tolerant fallbacks.
        desired = (sheet_name or '').strip()
        candidates: list[str] = []
        if desired:
            candidates.append(desired)
        # Common variations seen in LTP workbooks
        candidates.extend(['Planned', 'Planning', 'PLANING'])
        resolved_sheet = ''
        for cand in candidates:
            if cand in wb.sheetnames:
                resolved_sheet = cand
                break
            lower_match = next((s for s in wb.sheetnames if s.lower() == cand.lower()), '')
            if lower_match:
                resolved_sheet = lower_match
                break

        if not resolved_sheet:
            print(f"[LTP] Sheet '{sheet_name}' not found in {workbook_path.name} (available: {wb.sheetnames})")
            return planned

        ws = wb[resolved_sheet]
        start_col_idx = column_index_from_string(str(date_start_col).strip() or 'X')
        end_col_idx = column_index_from_string(str(date_end_col).strip() or 'BW')
        if end_col_idx < start_col_idx:
            start_col_idx, end_col_idx = end_col_idx, start_col_idx

        # Date headers: use configured row, but auto-detect if it doesn't look right.
        date_headers: list[date | None] = []
        for col in range(start_col_idx, end_col_idx + 1):
            date_headers.append(_coerce_header_to_date(ws.cell(row=date_row, column=col).value))

        if sum(1 for d in date_headers if d) < 3:
            scan_max = min(ws.max_row, 200)
            best_row = None
            best_count = 0
            for r in range(1, scan_max + 1):
                cnt = 0
                for col in range(start_col_idx, end_col_idx + 1):
                    if _coerce_header_to_date(ws.cell(row=r, column=col).value):
                        cnt += 1
                if cnt > best_count:
                    best_count = cnt
                    best_row = r

            if best_row is not None and best_count >= 3:
                print(f"[LTP] Date header row auto-detected: {best_row} ({best_count} date-like columns)")
                date_row = best_row
                date_headers = []
                for col in range(start_col_idx, end_col_idx + 1):
                    date_headers.append(_coerce_header_to_date(ws.cell(row=date_row, column=col).value))

        ref_triplet, ref_pair = _load_ltp_reference(ref_csv)
        target_label = (label or '').strip().lower()

        # Try to find the label column; if not found, proceed without label filtering.
        label_col_idx: int | None = 4 if target_label else None
        if target_label:
            scan_rows = min(ws.max_row, 500)
            direct_hits = 0
            for r in range(1, scan_rows + 1):
                v = ws.cell(row=r, column=4).value
                if str(v or '').strip().lower() == target_label:
                    direct_hits += 1
            if direct_hits == 0:
                best_col = None
                best_hits = 0
                scan_cols = min(ws.max_column, 15)
                for c in range(1, scan_cols + 1):
                    hits = 0
                    for r in range(1, scan_rows + 1):
                        v = ws.cell(row=r, column=c).value
                        if str(v or '').strip().lower() == target_label:
                            hits += 1
                    if hits > best_hits:
                        best_hits = hits
                        best_col = c
                if best_col is not None and best_hits > 0:
                    label_col_idx = best_col
                    print(f"[LTP] Label column auto-detected: {label_col_idx} ({best_hits} matches)")
                else:
                    label_col_idx = None
                    print(f"[LTP] ERROR: Label '{label}' not found in sheet '{resolved_sheet}'; falling back to CSV")
                    return load_planned_from_ltp_csv(fallback_csv) if fallback_csv else planned

        for row_idx in range(1, ws.max_row + 1):
            if label_col_idx is not None:
                label_raw = ws.cell(row=row_idx, column=label_col_idx).value
                if (str(label_raw or '').strip().lower()) != target_label:
                    continue

            project_raw = ws.cell(row=row_idx, column=1).value
            model_raw = ws.cell(row=row_idx, column=2).value
            project_key = _normalize_ltp_key(project_raw)
            model_key = _normalize_ltp_key(model_raw)
            if not project_key or not model_key:
                continue

            row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=3).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=5).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=6).value)
            if not row_type:
                row_type = _normalize_ltp_type(ws.cell(row=row_idx, column=4).value)
            if not row_type:
                row_type = _normalize_ltp_type(project_key) or _normalize_ltp_type(model_key)
            if not row_type:
                # Try color inference in schedule region first, then broaden to include early columns.
                row_type = _infer_type_from_row(ws, row_idx, start_col_idx, end_col_idx)
            if not row_type:
                row_type = _infer_type_from_row(ws, row_idx, 1, max(end_col_idx, 10))
            if not row_type:
                row_type = _infer_type_from_context(ws, row_idx)

            if row_type:
                codes = ref_triplet.get((project_key, model_key, row_type), [])
            else:
                codes = ref_pair.get((project_key, model_key), [])

            # If we inferred a type but the typed mapping doesn't exist, try safe fallbacks:
            # - opposite type (in case inference is wrong)
            # - unique pair mapping (when only one prod line exists for this project/model)
            if row_type and not codes:
                other = 'ASSY' if row_type == 'SEW' else 'SEW'
                other_codes = ref_triplet.get((project_key, model_key, other), [])
                if other_codes:
                    codes = other_codes
                    print(
                        f"[LTP] INFO: Using {other} mapping for {project_key}/{model_key} (inferred {row_type})"
                    )
                else:
                    pair_codes = ref_pair.get((project_key, model_key), [])
                    uniq_pair = sorted(set(norm_code(c) for c in pair_codes if norm_code(c)))
                    if len(uniq_pair) == 1:
                        codes = uniq_pair
                        print(
                            f"[LTP] INFO: Using unique pair mapping for {project_key}/{model_key} (inferred {row_type}) -> {codes[0]}"
                        )

            # If multiple prod lines match, allocate to all of them (sum accordingly)
            codes = [norm_code(c) for c in codes if norm_code(c)]
            codes = sorted(set(codes))
            if len(codes) > 1 and not row_type:
                print(f"[LTP] WARNING: Multiple prod lines for {project_key}/{model_key} (UNKNOWN type); row skipped")
                continue
            if len(codes) > 1:
                print(f"[LTP] INFO: Multiple prod lines for {project_key}/{model_key} ({row_type}): {codes}")

            if not codes:
                print(f"[LTP] WARNING: No prod line mapping for {project_key}/{model_key} ({row_type or 'UNKNOWN'})")
                continue

            per_day: dict[date, int] = {}
            for offset, col in enumerate(range(start_col_idx, end_col_idx + 1)):
                d = date_headers[offset] if offset < len(date_headers) else None
                if not d:
                    continue
                val = ws.cell(row=row_idx, column=col).value
                try:
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        qty = 0
                    else:
                        qty = int(round(float(val)))
                except Exception:
                    qty = 0
                if qty:
                    per_day[d] = per_day.get(d, 0) + qty

            if not per_day:
                continue

            for code in codes:
                planned.setdefault(code, {})
                for d, qty in per_day.items():
                    planned[code][d] = planned[code].get(d, 0) + qty
    finally:
        try:
            wb.close()
        except Exception:
            pass

    print(f"[LTP] Loaded {len(planned)} production lines from LTP workbook")
    return planned


def fetch_produced_by_day(start_d: date, end_d: date):
    """Return dict: {line_code: {date: qty_float}} for [start_d, end_d]."""
    data: dict[str, dict[date, float]] = {}
    conn = get_db_connection()
    try:
        cols = _get_tr_hist_columns(conn)
        has_tr_prod_line = 'tr_prod_line' in cols

        if has_tr_prod_line:
            line_expr = "COALESCE(NULLIF(LTRIM(RTRIM(tr.tr_prod_line)),''), NULLIF(LTRIM(RTRIM(pt.pt_prod_line)),''))"
        else:
            line_expr = "pt.pt_prod_line"

        tr_types = PVS_PROD_TR_TYPES or ['RCT-WO']
        placeholders = ','.join('?' for _ in tr_types)
        sql = (
            f"SELECT {line_expr} AS line, CAST(tr.tr_effdate AS date) AS d, "
            "SUM(CAST(tr.tr_qty_loc AS DECIMAL(18,2))) AS qty "
            "FROM dbo.tr_hist tr "
            "LEFT JOIN dbo.pt_mstr pt ON tr.tr_part = pt.pt_part "
            f"WHERE UPPER(LTRIM(RTRIM(tr.tr_type))) IN ({placeholders}) "
            "AND tr.tr_effdate >= ? AND tr.tr_effdate < DATEADD(day, 1, ?) "
            "AND tr.tr_qty_loc > 0 "
            f"GROUP BY {line_expr}, CAST(tr.tr_effdate AS date)"
        )

        cur = conn.cursor()
        cur.execute(sql, *tr_types, start_d, end_d)
        for line_code, d, qty in cur.fetchall():
            if not line_code:
                continue
            code = norm_code(line_code)
            data.setdefault(code, {})[d] = float(qty or 0)
        cur.close()
    finally:
        conn.close()
    print(f"[DB] Produced rows loaded: {len(data)} lines (range {start_d}..{end_d})")
    return data


def daterange(d0: date, d1: date):
    d = d0
    while d <= d1:
        yield d
        d += timedelta(days=1)


def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())  # Monday


def aggregate(values_by_day: dict[date, float], start_d: date, end_d: date) -> float:
    total = 0.0
    for d in daterange(start_d, end_d):
        total += float(values_by_day.get(d, 0) or 0)
    return total


def compute_metrics():
    today = date.today()
    wd = today.weekday()  # Monday=0 ... Sunday=6
    # Daily window:
    # - By default: yesterday only
    # - On Monday (optional): include Fri+Sat as "one day" window
    if wd in (6, 0) and PVS_SHOW_WEEKEND_ON_MONDAY:
        as_of = today - timedelta(days=1 if wd == 6 else 2)  # Saturday
        daily_start = as_of - timedelta(days=1)  # Friday
    else:
        as_of = today - timedelta(days=1)
        daily_start = as_of

    start_month = as_of.replace(day=1)
    start_week = monday_of_week(as_of)

    if PVS_REGENERATE_INPUTS:
        try:
            _run_production_sql_and_overwrite_csvs(PVS_PROD_SQL_PATH)
        except Exception as e:
            print(f"[SQL] WARNING: Production SQL regeneration failed: {e}")
        if PVS_EXPORT_LTP_REF_EXTRACT:
            try:
                _export_ltp_ref_extract_csv(
                    PVS_LTP_REF_EXTRACT_CSV,
                    PVS_LTP_DIR,
                    PVS_LTP_SHEET,
                    PVS_LTP_LABEL,
                    PVS_LTP_REF_CSV,
                    PVS_LTP_KEYWORDS,
                    PVS_LTP_DATE_ROW,
                    PVS_LTP_DATE_START_COL,
                    PVS_LTP_DATE_END_COL,
                )
            except Exception as e:
                print(f"[LTP-EXTRACT] WARNING: export failed: {e}")

    try:
        planned_pages = load_planned_pages_from_ltp(
            PVS_LTP_DIR,
            PVS_LTP_SHEET,
            PVS_LTP_LABEL,
            PVS_LTP_REF_CSV,
            PVS_LTP_KEYWORDS,
            PVS_LTP_DATE_ROW,
            PVS_LTP_DATE_START_COL,
            PVS_LTP_DATE_END_COL,
            PVS_LTP_WORKDAYS_PER_WEEK,
            PVS_LTP_FALLBACK_CSV,
        )

        master = _load_master_list(os.path.join(_BASE_DIR, 'PVS', 'master_list.csv'))
        planned_project = _canonicalize_series(planned_pages.get('PROJECT', {}), master.get('PROJECT') or [])
        planned_sew = _canonicalize_series(planned_pages.get('SEW', {}), master.get('SEW') or [])
        planned_assy = _canonicalize_series(planned_pages.get('ASSY', {}), master.get('ASSY') or [])

        _write_monthly_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Day', '1_PVS_per_Project.csv'),
            planned_project,
            start_month,
        )
        _write_monthly_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Day', '2_PVS_per_SEW.csv'),
            planned_sew,
            start_month,
        )
        _write_monthly_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Day', '3_PVS_per_ASSY.csv'),
            planned_assy,
            start_month,
        )

        _write_weekly_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Week', '1_PVS_per_Project.csv'),
            planned_project,
            start_month,
        )
        _write_weekly_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Week', '2_PVS_per_SEW.csv'),
            planned_sew,
            start_month,
        )
        _write_weekly_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Week', '3_PVS_per_ASSY.csv'),
            planned_assy,
            start_month,
        )

        _write_month_total_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Month', '1_PVS_per_Project.csv'),
            planned_project,
            start_month,
        )
        _write_month_total_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Month', '2_PVS_per_SEW.csv'),
            planned_sew,
            start_month,
        )
        _write_month_total_csv_by_label(
            os.path.join(_BASE_DIR, 'PVS', 'Planned', 'Month', '3_PVS_per_ASSY.csv'),
            planned_assy,
            start_month,
        )
    except Exception as e:
        print(f"[PVS] WARNING: Could not export planned CSVs: {e}")

    csv_res = _compute_metrics_from_page_csvs(as_of, daily_start, start_week, start_month)
    if csv_res is not None:
        return csv_res

    print(f"[COMPUTE] Computing metrics for date: {as_of}")
    print(
        f"[COMPUTE] PVS_PLAN_SOURCE={PVS_PLAN_SOURCE}, PVS_USE_WH_RECEIPT={PVS_USE_WH_RECEIPT}, "
        f"PVS_RECALC_XLSX={PVS_RECALC_XLSX}, PVS_PLANNED_XLSX={PVS_PLANNED_XLSX}"
    )

    if PVS_PLAN_SOURCE == 'ltp_formulas':
        print("[COMPUTE] Using LTP_formulas.xlsx for planned schedule...")
        planned = load_planned_from_ltp_formulas_xlsx(PVS_LTP_FORMULAS_XLSX, PVS_LTP_FORMULAS_SHEET)
    elif PVS_PLAN_SOURCE == 'ltp':
        print("[COMPUTE] Using LTP workbook for planned schedule...")
        planned = load_planned_from_ltp(
            PVS_LTP_DIR,
            PVS_LTP_SHEET,
            PVS_LTP_LABEL,
            PVS_LTP_REF_CSV,
            PVS_LTP_KEYWORDS,
            PVS_LTP_DATE_ROW,
            PVS_LTP_DATE_START_COL,
            PVS_LTP_DATE_END_COL,
            PVS_LTP_FALLBACK_CSV,
        )
    elif PVS_PLAN_SOURCE == 'ltp_csv':
        print("[COMPUTE] Using LTP fallback CSV for planned schedule...")
        planned = load_planned_from_ltp_csv(PVS_LTP_FALLBACK_CSV)
    elif PVS_USE_WH_RECEIPT:
        print("[COMPUTE] Using WH Receipt workbook for planned schedule (no Excel COM)...")
        planned = load_planned_from_wh_receipt(
            PVS_EXTERNAL_XLSX,
            PVS_EXTERNAL_SHEET,
            PVS_EXTERNAL_TARGET_LABEL,
        )
    else:
        if PVS_RECALC_XLSX:
            print("[COMPUTE] Attempting Excel recalculation of Planned_qtys.xlsx...")
            recalc_excel_workbook(PVS_PLANNED_XLSX)
        else:
            print("[COMPUTE] Excel recalc disabled - using cached values in Planned_qtys.xlsx")

        planned = load_planned_xlsx(PVS_PLANNED_XLSX)

    # LTP sources are weekly buckets (Monday dates). Expand to daily so Daily/WTD/MTD schedules work.
    if PVS_PLAN_SOURCE in ('ltp', 'ltp_csv', 'ltp_formulas') and _ltp_looks_weekly(planned):
        print(f"[COMPUTE] Expanding weekly LTP plan to daily using {PVS_LTP_WORKDAYS_PER_WEEK} workdays/week")
        planned = _expand_weekly_plan_to_daily(planned, PVS_LTP_WORKDAYS_PER_WEEK)
    # Fetch production from the earliest window we need (daily/WTD/MTD) to avoid undercounting
    # when week spans a month boundary or when daily view uses Fri+Sat.
    produced_start = min(start_month, start_week, daily_start)
    produced = fetch_produced_by_day(produced_start, as_of)
    mapping = load_map_csv(PVS_MAP_CSV)
    ref_meta = _load_ref_meta(PVS_LTP_REF_CSV)

    # Monthly OLK targets by prod line (for OLK column and OLK adherence)
    # OLK targets: prefer PVS/OLK.csv (display labels); fall back to Monthly_OLK.xlsx
    raw_olk = load_olk_csv(PVS_OLK_CSV)
    if not raw_olk:
        raw_olk = load_monthly_olk(PVS_OLK_XLSX)
    olk_by_code: dict[str, float] = {}
    if raw_olk:
        # Build reverse map: display name -> code
        display_to_code: dict[str, str] = {}
        for code_key, disp_name in mapping.items():
            if not disp_name:
                continue
            display_to_code[disp_name.upper()] = code_key

        for code_key, m in ref_meta.items():
            for label in (m.get('sew'), m.get('assy'), m.get('project_group')):
                if label:
                    display_to_code[str(label).strip().upper()] = code_key

        code_keys = list(mapping.keys())

        for key, val in raw_olk.items():
            if not val:
                continue

            key_norm = (key or "").strip().upper()
            target_code: str | None = None

            # 1) Exact match on prod line code
            if key_norm in mapping:
                target_code = key_norm
            else:
                # 2) Exact match on display / project name
                target_code = display_to_code.get(key_norm)

            # 3) Fuzzy match: key contains a known code (e.g. "Z_FG - CDPO")
            if target_code is None:
                for code_key in code_keys:
                    if code_key in key_norm:
                        target_code = code_key
                        break

            # 4) Fuzzy match: key contains a known display name
            if target_code is None:
                for code_key, disp_name in mapping.items():
                    if not disp_name:
                        continue
                    if disp_name.upper() in key_norm:
                        target_code = code_key
                        break

            if target_code is not None:
                olk_by_code[target_code] = olk_by_code.get(target_code, 0.0) + float(val or 0.0)
            else:
                # Fallback: keep under original key so at least it's not lost
                olk_by_code[key_norm] = olk_by_code.get(key_norm, 0.0) + float(val or 0.0)

        def _base_project_name(name: str) -> str:
            s = (name or '').strip().upper()
            for suf in (' - SEW', '- SEW', ' - ASSY', '- ASSY'):
                if s.endswith(suf):
                    s = s[:-len(suf)]
                    break
            s = s.strip()
            while s.endswith('-'):
                s = s[:-1].strip()
            return s

        pairs: dict[str, dict[str, str]] = {}
        for code_key, disp_name in mapping.items():
            if not disp_name:
                continue
            disp_u = disp_name.strip().upper()
            if 'SEW' in disp_u:
                base = _base_project_name(disp_u)
                pairs.setdefault(base, {})['SEW'] = code_key
            elif 'ASSY' in disp_u:
                base = _base_project_name(disp_u)
                pairs.setdefault(base, {})['ASSY'] = code_key

        for pair in pairs.values():
            sew_code = pair.get('SEW')
            assy_code = pair.get('ASSY')
            if not sew_code or not assy_code:
                continue
            sew_val = float(olk_by_code.get(sew_code, 0.0) or 0.0)
            assy_val = float(olk_by_code.get(assy_code, 0.0) or 0.0)
            if sew_val <= 0 and assy_val > 0:
                olk_by_code[sew_code] = assy_val
            elif assy_val <= 0 and sew_val > 0:
                olk_by_code[assy_code] = sew_val
    else:
        olk_by_code = {}

    # Build union of all line codes seen in plan or production
    codes = sorted(set(planned.keys()) | set(produced.keys()))

    rows = []
    for code in codes:
        m = ref_meta.get(code, {})
        t = (m.get('type') or '').strip().upper()
        if t == 'SEW':
            disp = m.get('sew') or mapping.get(code, code)
        elif t == 'ASSY':
            disp = m.get('assy') or mapping.get(code, code)
        else:
            disp = mapping.get(code, code)
        plan_days = planned.get(code, {})
        prod_days = produced.get(code, {})

        # Plans (previous business day or Fri+Sat)
        plan_day = int(aggregate(plan_days, daily_start, as_of))
        plan_wtd = int(aggregate(plan_days, start_week, as_of))
        plan_mtd = int(aggregate(plan_days, start_month, as_of))

        # Produced (previous business day or Fri+Sat)
        prod_day = float(aggregate(prod_days, daily_start, as_of))
        prod_wtd = float(aggregate(prod_days, start_week, as_of))
        prod_mtd = float(aggregate(prod_days, start_month, as_of))

        # Deltas
        d_day = prod_day - plan_day
        d_wtd = prod_wtd - plan_wtd
        d_mtd = prod_mtd - plan_mtd

        def adherence(delta: float, schedule: float) -> float | None:
            if schedule <= 0:
                return None
            pct = (delta / schedule) * 100.0
            # Clamp extreme outliers instead of zeroing them
            if pct > PVS_ADHERENCE_CLAMP:
                return PVS_ADHERENCE_CLAMP
            if pct < -PVS_ADHERENCE_CLAMP:
                return -PVS_ADHERENCE_CLAMP
            return pct

        # Determine category (SEW, ASSY, or OTHER)
        if t in ('SEW', 'ASSY'):
            category = t
        else:
            disp_upper = str(disp).upper()
            normalized_name = disp_upper.replace('-', ' ').strip()
            if normalized_name in SEW_NAME_OVERRIDES or 'SEW' in disp_upper:
                category = 'SEW'
            elif 'ASSY' in disp_upper:
                category = 'ASSY'
            else:
                category = 'OTHER'

        # Monthly OLK qty for this prod line (whole-month target) and OLK adherence (ratio)
        line_olk = float(olk_by_code.get(code, 0.0) or 0.0)
        if line_olk > 0:
            adh_olk_pct = (prod_mtd / line_olk) * 100.0
        else:
            adh_olk_pct = 0.0

        adh_mtd = adherence(d_mtd, plan_mtd)
        adh_wtd = adherence(d_wtd, plan_wtd)
        adh_day = adherence(d_day, plan_day)

        rows.append({
            'code': code,
            'line': disp,
            'category': category,
            'mtd': {
                'olk': int(round(line_olk)),
                'adh_olk_pct': round(adh_olk_pct, 1),
                'schedule': plan_mtd,
                'production': round(prod_mtd, 2),
                'delta': round(d_mtd, 2),
                # Use cumulative planned as denominator for MTD adherence
                'adherence_pct': round(adh_mtd, 1) if adh_mtd is not None else None,
            },
            'wtd': {
                'schedule': plan_wtd,
                'production': round(prod_wtd, 2),
                'delta': round(d_wtd, 2),
                # Use cumulative planned as denominator for WTD adherence
                'adherence_pct': round(adh_wtd, 1) if adh_wtd is not None else None,
            },
            'daily': {
                'schedule': plan_day,
                'production': round(prod_day, 2),
                'delta': round(d_day, 2),
                'adherence_pct': round(adh_day, 1) if adh_day is not None else None,
            }
        })

    # Sort: SEW first, then ASSY, then OTHER; within each category sort by line name
    category_order = {'SEW': 0, 'ASSY': 1, 'OTHER': 2}
    rows.sort(key=lambda r: (category_order.get(r['category'], 2), r['line']))

    # Compute totals for pie charts (MTD, WTD, Daily) per category
    def _empty_bucket():
        return {'schedule': 0, 'production': 0}

    totals = {
        'sew': {
            'mtd': _empty_bucket(),
            'wtd': _empty_bucket(),
            'daily': _empty_bucket(),
        },
        'assy': {
            'mtd': _empty_bucket(),
            'wtd': _empty_bucket(),
            'daily': _empty_bucket(),
        },
        'other': {
            'mtd': _empty_bucket(),
            'wtd': _empty_bucket(),
            'daily': _empty_bucket(),
        },
        'all': {
            'mtd': _empty_bucket(),
            'wtd': _empty_bucket(),
            'daily': _empty_bucket(),
        },
    }

    # OLK totals per category (for OLK pies on SEW/ASSY pages, MTD only)
    olk_totals = {
        'sew': {'olk': 0.0, 'production': 0.0},
        'assy': {'olk': 0.0, 'production': 0.0},
    }

    def _accumulate(bucket: dict, key: str, sched: int | float, prod: int | float) -> None:
        b = bucket[key]
        b['schedule'] += int(sched or 0)
        b['production'] += float(prod or 0.0)

    # Buckets for grouped totals (pairs, CV, Nissan, singles)
    def _empty_group_window() -> dict[str, float]:
        return {
            'schedule': 0.0,
            'production': 0.0,
            'delta': 0.0,
            'adherence_pct': 0.0,
            'olk': 0.0,
            'olk_projects': {},
        }

    group_buckets: dict[str, dict[str, dict[str, float]]] = {}

    def _group_project_key(disp_name: str) -> str:
        s = (disp_name or '').strip().upper()
        if ' - ' in s:
            s = s.split(' - ')[0].strip()
        for suf in (' - SEW', '- SEW', ' - ASSY', '- ASSY'):
            if s.endswith(suf):
                s = s[:-len(suf)].strip()
                break
        while s.endswith('-'):
            s = s[:-1].strip()
        return s

    for r in rows:
        cat = r['category']

        # All categories combined
        _accumulate(totals['all'], 'mtd', r['mtd']['schedule'], r['mtd']['production'])
        _accumulate(totals['all'], 'wtd', r['wtd']['schedule'], r['wtd']['production'])
        _accumulate(totals['all'], 'daily', r['daily']['schedule'], r['daily']['production'])

        # Category-specific buckets
        if cat == 'SEW':
            bucket = totals['sew']
            olk_key = 'sew'
        elif cat == 'ASSY':
            bucket = totals['assy']
            olk_key = 'assy'
        else:
            bucket = totals['other']
            olk_key = None

        _accumulate(bucket, 'mtd', r['mtd']['schedule'], r['mtd']['production'])
        _accumulate(bucket, 'wtd', r['wtd']['schedule'], r['wtd']['production'])
        _accumulate(bucket, 'daily', r['daily']['schedule'], r['daily']['production'])

        # Accumulate OLK totals per category (MTD only)
        if olk_key is not None:
            line_olk = float(r['mtd'].get('olk', 0) or 0.0)
            if line_olk:
                cat_olk = olk_totals[olk_key]
                cat_olk['olk'] += line_olk
                cat_olk['production'] += float(r['mtd'].get('production', 0.0) or 0.0)

        # Grouped totals logic
        code = r.get('code', '')
        m = ref_meta.get(str(code), {}) if code else {}
        group_name = (m.get('project_group') or '').strip()
        if not group_name:
            disp = r['line']
            disp_upper = disp.upper()
            base = disp.split(' - ')[0].strip().upper() if ' - ' in disp else disp_upper
            group_name = GROUP_OVERRIDES.get(base)
            if not group_name:
                # If no override, for SEW/ASSY pairs use base (e.g. CDPO, JLR, ...)
                if ' - ' in disp and (cat in ('SEW', 'ASSY')):
                    group_name = base
                else:
                    # Single line group (BJA, FIAT, etc.)
                    group_name = base

        grp = group_buckets.setdefault(group_name, {
            'mtd': _empty_group_window(),
            'wtd': _empty_group_window(),
            'daily': _empty_group_window(),
        })

        # Accumulate group OLK on MTD window
        proj_key = _group_project_key(disp)
        line_olk = float(r['mtd'].get('olk', 0) or 0.0)
        if line_olk:
            olk_projects = grp['mtd'].setdefault('olk_projects', {})
            prev = float(olk_projects.get(proj_key, 0.0) or 0.0)
            if line_olk > prev:
                olk_projects[proj_key] = line_olk

        for key in ('mtd', 'wtd', 'daily'):
            win = grp[key]
            src = r[key]
            win['schedule'] += float(src['schedule'] or 0.0)
            win['production'] += float(src['production'] or 0.0)

    # Finalize grouped totals: compute delta and adherence for each window
    group_totals: list[dict[str, object]] = []
    for name in sorted(group_buckets.keys()):
        src = group_buckets[name]
        agg: dict[str, object] = {'group': name}
        for key in ('mtd', 'wtd', 'daily'):
            win = src[key]
            sched = float(win['schedule'])
            prod = float(win['production'])
            delta = prod - sched
            adh = adherence(delta, sched)
            win_out = {
                'schedule': int(round(sched)),
                'production': round(prod, 2),
                'delta': round(delta, 2),
                'adherence_pct': round(adh, 1) if adh is not None else None,
            }
            if key == 'mtd':
                olk_projects = win.get('olk_projects')
                if isinstance(olk_projects, dict) and olk_projects:
                    olk_qty = sum(float(v or 0.0) for v in olk_projects.values())
                else:
                    olk_qty = float(win.get('olk', 0.0) or 0.0)
                if olk_qty > 0:
                    adh_olk_pct = (prod / olk_qty) * 100.0
                else:
                    adh_olk_pct = 0.0
                win_out['olk'] = int(round(olk_qty))
                win_out['adh_olk_pct'] = round(adh_olk_pct, 1)
            agg[key] = win_out
        group_totals.append(agg)

    # Compute category-level OLK totals for SEW/ASSY (for OLK pies)
    olk_totals_out: dict[str, dict[str, float]] = {}
    for key in ('sew', 'assy'):
        src = olk_totals.get(key, {})
        olk_qty = float(src.get('olk', 0.0) or 0.0)
        prod_qty = float(src.get('production', 0.0) or 0.0)
        if olk_qty > 0:
            adh_olk_pct = (prod_qty / olk_qty) * 100.0
        else:
            adh_olk_pct = 0.0
        olk_totals_out[key] = {
            'olk': int(round(olk_qty)),
            'production': round(prod_qty, 2),
            'adh_olk_pct': round(adh_olk_pct, 1),
        }

    return {
        'success': True,
        'date': as_of.strftime('%Y-%m-%d'),
        'rows': rows,
        'totals': totals,
        'group_totals': group_totals,
        'olk_totals': olk_totals_out,
    }


@app.route('/')
def index():
    return render_template('pvs.html', version=str(int(datetime.now().timestamp())))


@app.route('/api/pvs')
def api_pvs():
    try:
        return jsonify(compute_metrics())
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/health')
def health():
    return jsonify({'status': 'healthy', 'ts': datetime.now().isoformat()})


if __name__ == '__main__':
    from waitress import serve
    print('=' * 70)
    print(f'Running PVS app at http://{FLASK_HOST}:{PVS_PORT}')
    print('=' * 70)
    serve(app, host=FLASK_HOST, port=PVS_PORT, threads=4)
